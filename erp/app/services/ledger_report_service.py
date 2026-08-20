"""Reports & Analysis → Ledger Report (search + invoice-style preview)."""

from __future__ import annotations

import csv
import io
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from xml.dom import minidom

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text

from app.extensions import db
from app.services.ledger_export_service import LedgerExportService


def _iso(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


class LedgerReportService:
    """Ledger search and preview for bank / customer / work-category / item."""

    KINDS = ("bank", "customer", "work", "item")

    @staticmethod
    def _money(value) -> Decimal:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))

    @staticmethod
    def _parse_date(raw: str | None, fallback: date) -> date:
        value = (raw or "").strip()
        if not value:
            return fallback
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return fallback

    @staticmethod
    def _resolve_period(
        date_from: date | None, date_to: date | None
    ) -> tuple[date, date]:
        today = date.today()
        return date_from or date(2000, 1, 1), date_to or today

    def _sort_date_value(self, date_str: str) -> str:
        raw = (date_str or "").strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw, fmt).date().isoformat()
            except ValueError:
                continue
        return raw

    def _decorate_line(self, line: dict[str, Any], *, link: dict[str, Any] | None = None) -> dict[str, Any]:
        line["sort_date"] = self._sort_date_value(str(line.get("date") or ""))
        kind = (line.get("kind") or "txn").strip()
        if kind != "txn":
            line["can_edit"] = False
            line["can_delete"] = False
            line["source_url"] = ""
            line["source_module"] = ""
            line["source_module_id"] = ""
            line["work_type"] = ""
            return line
        link = link or {}
        module = (link.get("source_module") or "").strip()
        mid = link.get("source_module_id")
        line["can_edit"] = bool(link.get("can_open") and link.get("source_url"))
        line["can_delete"] = bool(module and mid)
        line["source_url"] = link.get("source_url") or ""
        line["source_module"] = module
        line["source_module_id"] = mid or ""
        line["work_type"] = (link.get("work_type") or "").strip()
        return line

    def _period_fields(self, date_from: date | None, date_to: date | None) -> dict[str, Any]:
        return {
            "date_from": date_from,
            "date_to": date_to,
            "date_from_iso": _iso(date_from),
            "date_to_iso": _iso(date_to),
        }

    def _dash(self):
        from app.services.dashboard_service import DashboardService

        return DashboardService()

    @staticmethod
    def _fy_start(today: date | None = None) -> date:
        today = today or date.today()
        year = today.year if today.month >= 4 else today.year - 1
        return date(year, 4, 1)

    def search_ledgers(
        self,
        *,
        kind: str | None = None,
        search: str | None = None,
        limit: int = 100,
    ) -> list[dict[str, Any]]:
        kind_key = (kind or "all").strip().lower()
        needle = (search or "").strip()
        lim = max(1, min(int(limit or 100), 300))
        rows: list[dict[str, Any]] = []

        if kind_key in ("all", "bank"):
            rows.extend(self._search_banks(needle, lim if kind_key == "bank" else min(lim, 50)))
        if kind_key in ("all", "customer"):
            if kind_key == "customer" or len(needle) >= 2:
                rows.extend(
                    self._search_customers(needle, lim if kind_key == "customer" else min(lim, 50))
                )
        if kind_key in ("all", "work"):
            rows.extend(self._search_works(needle, lim if kind_key == "work" else min(lim, 50)))
        if kind_key in ("all", "item"):
            rows.extend(self._search_items(needle, lim if kind_key == "item" else min(lim, 50)))

        return rows[:lim]

    def _search_banks(self, search: str, limit: int) -> list[dict[str, Any]]:
        export = LedgerExportService()
        banks = export.list_bank_accounts(search=search or None)
        result = []
        for row in banks[:limit]:
            result.append(
                {
                    "kind": "bank",
                    "id": row["account_id"],
                    "label": row["label"],
                    "subtitle": row.get("account_holder") or row.get("account_type") or "Bank Account",
                    "meta": row.get("masked_account") or "",
                    "txn_count": int(row.get("txn_count") or 0),
                    "active": bool(row.get("active", True)),
                }
            )
        return result

    def _search_customers(self, search: str, limit: int) -> list[dict[str, Any]]:
        if len((search or "").strip()) < 2:
            return []
        export = LedgerExportService()
        customers = export.list_customers(search=search, limit=limit)
        result = []
        for row in customers:
            bits = [b for b in (row.get("mobile_number"), row.get("pan_number")) if b]
            result.append(
                {
                    "kind": "customer",
                    "id": row["customer_id"],
                    "label": row["customer_name"],
                    "subtitle": " · ".join(bits) if bits else "Customer",
                    "meta": row.get("status") or "",
                    "txn_count": int(row.get("txn_count") or 0),
                    "active": (row.get("status") or "Active").strip().lower() == "active",
                }
            )
        return result

    def _search_works(self, search: str, limit: int) -> list[dict[str, Any]]:
        params: dict[str, Any] = {"lim": limit}
        search_sql = ""
        needle = (search or "").strip()
        if needle:
            params["like"] = f"%{needle}%"
            search_sql = """
              AND (
                w.WorkName LIKE :like
                OR w.LedgerKind LIKE :like
              )
            """
        rows = db.session.execute(
            text(
                f"""
                SELECT TOP (:lim)
                    w.WorkID,
                    w.WorkName,
                    w.LedgerKind,
                    ISNULL(w.ActiveStatus, 1) AS ActiveStatus,
                    (
                        SELECT COUNT(1)
                        FROM dbo.JTCSDailyTransaction d
                        WHERE d.Status = N'Posted'
                          AND (
                            d.WorkType = w.WorkName
                            OR d.SubWorkType = w.WorkName
                            OR d.SubWorkType LIKE N'%' + w.WorkName + N'%'
                          )
                    ) AS txn_count
                FROM dbo.WorkMaster w
                WHERE 1 = 1
                  {search_sql}
                ORDER BY
                    CASE WHEN ISNULL(w.ActiveStatus, 1) = 1 THEN 0 ELSE 1 END,
                    w.WorkName,
                    w.WorkID
                """
            ),
            params,
        ).mappings().all()
        return [
            {
                "kind": "work",
                "id": int(row["WorkID"]),
                "label": (row["WorkName"] or "").strip(),
                "subtitle": f"Work / Category · {(row['LedgerKind'] or '').strip() or '—'}",
                "meta": (row["LedgerKind"] or "").strip(),
                "txn_count": int(row["txn_count"] or 0),
                "active": bool(row["ActiveStatus"]),
            }
            for row in rows
        ]

    def _search_items(self, search: str, limit: int) -> list[dict[str, Any]]:
        params: dict[str, Any] = {"lim": limit}
        search_sql = ""
        needle = (search or "").strip()
        if needle:
            params["like"] = f"%{needle}%"
            search_sql = """
              AND (
                i.ItemCode LIKE :like
                OR i.ItemName LIKE :like
                OR ISNULL(i.HsnSac, N'') LIKE :like
              )
            """
        rows = db.session.execute(
            text(
                f"""
                SELECT TOP (:lim)
                    i.ItemID,
                    i.ItemCode,
                    i.ItemName,
                    i.HsnSac,
                    ISNULL(i.IsActive, 1) AS IsActive,
                    (
                        SELECT COUNT(1)
                        FROM dbo.GstInvoiceLine l
                        WHERE l.ItemID = i.ItemID
                    ) AS txn_count
                FROM dbo.ItemMaster i
                WHERE 1 = 1
                  {search_sql}
                ORDER BY
                    CASE WHEN ISNULL(i.IsActive, 1) = 1 THEN 0 ELSE 1 END,
                    i.OrderNo,
                    i.ItemName,
                    i.ItemID
                """
            ),
            params,
        ).mappings().all()
        return [
            {
                "kind": "item",
                "id": int(row["ItemID"]),
                "label": (row["ItemName"] or "").strip(),
                "subtitle": f"Item · {(row['ItemCode'] or '').strip()}"
                + (f" · HSN/SAC {(row['HsnSac'] or '').strip()}" if row["HsnSac"] else ""),
                "meta": (row["ItemCode"] or "").strip(),
                "txn_count": int(row["txn_count"] or 0),
                "active": bool(row["IsActive"]),
            }
            for row in rows
        ]

    def preview_ledger(
        self,
        kind: str,
        entity_id: int,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> dict[str, Any]:
        kind_key = (kind or "").strip().lower()
        if kind_key not in self.KINDS:
            raise ValueError("Invalid ledger type.")
        if kind_key == "bank":
            return self._simplify_export_ledger(
                LedgerExportService().bank_ledger_preview_data(
                    entity_id, date_from=date_from, date_to=date_to
                ),
                title="Bank Account Ledger",
            )
        if kind_key == "customer":
            return self._simplify_export_ledger(
                LedgerExportService().customer_ledger_preview_data(
                    entity_id, date_from=date_from, date_to=date_to
                ),
                title="Customer Ledger",
            )
        if kind_key == "work":
            return self._work_ledger_data(entity_id, date_from=date_from, date_to=date_to)
        return self._item_ledger_data(entity_id, date_from=date_from, date_to=date_to)

    def _prefetch_daily_source_rows(self, record_ids: list[int]) -> dict[int, dict[str, Any]]:
        ids = []
        seen: set[int] = set()
        for raw in record_ids:
            try:
                tid = int(raw)
            except (TypeError, ValueError):
                continue
            if tid <= 0 or tid in seen:
                continue
            seen.add(tid)
            ids.append(tid)
        if not ids:
            return {}
        result: dict[int, dict[str, Any]] = {}
        chunk_size = 400
        for start in range(0, len(ids), chunk_size):
            chunk = ids[start : start + chunk_size]
            placeholders = ", ".join(str(tid) for tid in chunk)
            rows = db.session.execute(
                text(
                    f"""
                    SELECT TransactionID, WorkType, SubWorkType, StampID, ReferenceNo
                    FROM JTCSDailyTransaction
                    WHERE TransactionID IN ({placeholders})
                    """
                )
            ).mappings().all()
            for row in rows:
                result[int(row["TransactionID"])] = dict(row)
        return result

    def _simplify_export_ledger(self, data: dict[str, Any], *, title: str) -> dict[str, Any]:
        """Map admin export ledger rows to Date / Description / Debit / Credit / Closing Balance."""
        dash = self._dash()
        kind_key = (data.get("kind") or "").strip().lower()
        daily_map: dict[int, dict[str, Any]] = {}
        if kind_key == "bank":
            daily_map = self._prefetch_daily_source_rows(
                [
                    line.get("source_record_id")
                    for line in (data.get("lines") or [])
                    if (line.get("kind") or "txn") == "txn"
                    and (line.get("source") or "").strip().lower()
                    in {"", "jtcsdailytransaction", "shcil"}
                    and line.get("source_record_id")
                ]
            )
        lines: list[dict[str, Any]] = []
        for line in data.get("lines") or []:
            kind = (line.get("kind") or "txn").strip()
            if kind in ("day_total", "month_total", "desc_total"):
                continue
            desc_parts = []
            if line.get("description"):
                desc_parts.append(str(line["description"]).strip())
            if line.get("bill"):
                desc_parts.append(f"Ref: {str(line['bill']).strip()}")
            if line.get("work"):
                desc_parts.append(str(line["work"]).strip())
            if line.get("reference") and kind == "txn":
                desc_parts.append(str(line["reference"]).strip())
            out = {
                "date": line.get("date") or "",
                "description": " · ".join([p for p in desc_parts if p]) or "—",
                "debit": line.get("debit"),
                "credit": line.get("credit"),
                "balance": line.get("balance"),
                "kind": kind,
            }
            link = None
            if kind == "txn":
                try:
                    if kind_key == "bank":
                        rec_id = line.get("source_record_id")
                        rec_int = int(rec_id) if rec_id else 0
                        daily = daily_map.get(rec_int)
                        if daily:
                            link = dash._source_link_for_daily(
                                transaction_id=int(daily["TransactionID"]),
                                work_type=daily.get("WorkType"),
                                sub_work_type=daily.get("SubWorkType"),
                                stamp_id=daily.get("StampID"),
                                reference=daily.get("ReferenceNo"),
                            )
                        else:
                            link = dash._source_link_for_bank_leg(
                                source_table=line.get("source"),
                                source_record_id=rec_int if rec_int else None,
                                bank_transaction_id=line.get("bank_transaction_id"),
                            )
                    elif kind_key == "customer":
                        txn_id = line.get("transaction_id")
                        stamp_id = line.get("stamp_id")
                        txn_int = int(txn_id) if txn_id not in (None, "") else 0
                        link = dash._source_link_for_daily(
                            transaction_id=txn_int if txn_int > 0 else None,
                            work_type=line.get("work_type"),
                            sub_work_type=line.get("sub_work_type"),
                            stamp_id=int(stamp_id) if stamp_id else None,
                            reference=line.get("bill") or line.get("reference"),
                        )
                        if not link.get("can_open") and line.get("obc_entry_id"):
                            link = dash._source_link_for_bank_cash_entry(int(line["obc_entry_id"]))
                except Exception:
                    link = None
            lines.append(self._decorate_line(out, link=link))
        period = self._period_fields(data.get("date_from"), data.get("date_to"))
        return {
            "kind": data.get("kind") or "ledger",
            "title": title,
            "entity_name": data.get("entity_name") or "",
            "entity_id": data.get("entity_id"),
            "meta": data.get("meta") or [],
            "headers": ["Date", "Description", "Debit", "Credit", "Closing Balance"],
            "lines": lines,
            "closing": data.get("closing") or Decimal("0.00"),
            **period,
        }

    def _work_ledger_data(
        self,
        work_id: int,
        *,
        date_from: date | None,
        date_to: date | None,
    ) -> dict[str, Any]:
        work = db.session.execute(
            text(
                """
                SELECT WorkID, WorkName, LedgerKind, ISNULL(ActiveStatus, 1) AS ActiveStatus
                FROM dbo.WorkMaster
                WHERE WorkID = :work_id
                """
            ),
            {"work_id": work_id},
        ).mappings().first()
        if work is None:
            raise ValueError("Work / Category not found.")

        date_from, date_to = self._resolve_period(date_from, date_to)
        work_name = (work["WorkName"] or "").strip()
        ledger_kind = (work["LedgerKind"] or "").strip().upper()

        # SQL Server forbids SUM( (SELECT SUM(...)) ); join payment totals instead.
        prior = db.session.execute(
            text(
                """
                SELECT
                    ISNULL(SUM(ISNULL(d.SaleAmount, 0) + ISNULL(d.IncomeAmount, 0)), 0) AS income_amt,
                    ISNULL(SUM(ISNULL(d.ExpenseAmount, 0)), 0) AS expense_amt,
                    ISNULL(SUM(ISNULL(pay.paid_amt, 0)), 0) AS paid_amt
                FROM dbo.JTCSDailyTransaction d
                LEFT JOIN (
                    SELECT TransactionID, SUM(Amount) AS paid_amt
                    FROM dbo.JTCSDailyTransactionPayment
                    GROUP BY TransactionID
                ) pay ON pay.TransactionID = d.TransactionID
                WHERE d.Status = N'Posted'
                  AND d.TransactionDate < :date_from
                  AND (
                    d.WorkType = :work_name
                    OR d.SubWorkType = :work_name
                    OR d.SubWorkType LIKE N'%' + :work_name + N'%'
                  )
                """
            ),
            {"work_name": work_name, "date_from": date_from},
        ).mappings().first()

        prior_income = self._money(prior["income_amt"] if prior else 0)
        prior_expense = self._money(prior["expense_amt"] if prior else 0)
        prior_paid = self._money(prior["paid_amt"] if prior else 0)
        # Income/Misc → credit nature; Expense → debit nature
        if ledger_kind.startswith("E"):
            opening = self._money(prior_expense - prior_paid)
        else:
            opening = self._money(prior_income - prior_paid)

        rows = db.session.execute(
            text(
                """
                SELECT
                    d.TransactionID, d.TransactionDate, d.WorkType, d.SubWorkType,
                    d.StampID, d.ReferenceNo, d.Description, d.Remarks,
                    ISNULL(d.SaleAmount, 0) AS SaleAmount,
                    ISNULL(d.IncomeAmount, 0) AS IncomeAmount,
                    ISNULL(d.ExpenseAmount, 0) AS ExpenseAmount,
                    (
                        SELECT ISNULL(SUM(p.Amount), 0)
                        FROM dbo.JTCSDailyTransactionPayment p
                        WHERE p.TransactionID = d.TransactionID
                    ) AS PaymentTotal
                FROM dbo.JTCSDailyTransaction d
                WHERE d.Status = N'Posted'
                  AND d.TransactionDate >= :date_from
                  AND d.TransactionDate <= :date_to
                  AND (
                    d.WorkType = :work_name
                    OR d.SubWorkType = :work_name
                    OR d.SubWorkType LIKE N'%' + :work_name + N'%'
                  )
                ORDER BY d.TransactionDate ASC, d.TransactionID ASC
                """
            ),
            {"work_name": work_name, "date_from": date_from, "date_to": date_to},
        ).mappings().all()

        lines: list[dict[str, Any]] = []
        running = opening
        lines.append(
            self._decorate_line(
                {
                    "date": date_from.strftime("%d/%m/%Y"),
                    "description": "Opening Balance",
                    "debit": Decimal("0.00"),
                    "credit": Decimal("0.00"),
                    "balance": running,
                    "kind": "opening",
                }
            )
        )

        dash = self._dash()
        for row in rows:
            income = self._money(row["SaleAmount"]) + self._money(row["IncomeAmount"])
            expense = self._money(row["ExpenseAmount"])
            payment = self._money(row["PaymentTotal"])
            if ledger_kind.startswith("E"):
                debit = expense
                credit = payment
            else:
                debit = payment
                credit = income
            if debit == 0 and credit == 0:
                # Fallback: show whichever amount exists
                if income > 0:
                    credit = income
                elif expense > 0:
                    debit = expense
            running = self._money(running + debit - credit)
            work_label = (row["WorkType"] or "").strip()
            sub = (row["SubWorkType"] or "").strip()
            if sub:
                work_label = f"{work_label} / {sub}" if work_label else sub
            desc_bits = [
                (row["Description"] or row["Remarks"] or "").strip(),
                work_label,
                (row["ReferenceNo"] or "").strip() or f"TXN-{row['TransactionID']}",
            ]
            txn_date = row["TransactionDate"]
            lines.append(
                self._decorate_line(
                    {
                        "date": txn_date.strftime("%d/%m/%Y") if txn_date else "",
                        "description": " · ".join([b for b in desc_bits if b]) or "Transaction",
                        "debit": debit,
                        "credit": credit,
                        "balance": running,
                        "kind": "txn",
                    },
                    link=dash._source_link_for_daily(
                        transaction_id=row["TransactionID"],
                        work_type=row["WorkType"],
                        sub_work_type=row["SubWorkType"],
                        stamp_id=row["StampID"],
                        reference=row["ReferenceNo"],
                    ),
                )
            )

        period = self._period_fields(date_from, date_to)
        return {
            "kind": "work",
            "title": "Work / Category Ledger",
            "entity_name": work_name,
            "entity_id": work_id,
            "meta": [
                ("Work / Category", work_name),
                ("Ledger Kind", (work["LedgerKind"] or "").strip() or "—"),
                ("Period", f"{date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}"),
            ],
            "headers": ["Date", "Description", "Debit", "Credit", "Closing Balance"],
            "lines": lines,
            "closing": running,
            **period,
        }

    def _item_ledger_data(
        self,
        item_id: int,
        *,
        date_from: date | None,
        date_to: date | None,
    ) -> dict[str, Any]:
        item = db.session.execute(
            text(
                """
                SELECT
                    ItemID, ItemCode, ItemName, HsnSac, Unit,
                    ISNULL(OpeningBalance, 0) AS OpeningBalance,
                    OpeningBalanceDate,
                    ISNULL(IsActive, 1) AS IsActive
                FROM dbo.ItemMaster
                WHERE ItemID = :item_id
                """
            ),
            {"item_id": item_id},
        ).mappings().first()
        if item is None:
            raise ValueError("Item not found.")

        date_from, date_to = self._resolve_period(date_from, date_to)
        item_name = (item["ItemName"] or "").strip()
        item_code = (item["ItemCode"] or "").strip()

        opening = Decimal("0.00")
        ob_date = item["OpeningBalanceDate"]
        if ob_date is None or ob_date <= date_from:
            opening = self._money(item["OpeningBalance"])

        prior = db.session.execute(
            text(
                """
                SELECT ISNULL(SUM(ISNULL(l.TaxableValue, 0)), 0)
                FROM dbo.GstInvoiceLine l
                INNER JOIN dbo.GstInvoice inv ON inv.InvoiceID = l.InvoiceID
                WHERE l.ItemID = :item_id
                  AND inv.InvoiceDate < :date_from
                """
            ),
            {"item_id": item_id, "date_from": date_from},
        ).scalar()
        # Sales reduce stock value / credit the item ledger
        opening = self._money(opening - self._money(prior))

        rows = db.session.execute(
            text(
                """
                SELECT
                    inv.InvoiceID,
                    inv.InvoiceNo,
                    inv.InvoiceDate,
                    inv.CustomerName,
                    l.Particulars,
                    l.Qty,
                    l.Unit,
                    l.Rate,
                    ISNULL(l.TaxableValue, 0) AS TaxableValue,
                    ISNULL(l.DiscountAmount, 0) AS DiscountAmount
                FROM dbo.GstInvoiceLine l
                INNER JOIN dbo.GstInvoice inv ON inv.InvoiceID = l.InvoiceID
                WHERE l.ItemID = :item_id
                  AND inv.InvoiceDate >= :date_from
                  AND inv.InvoiceDate <= :date_to
                ORDER BY inv.InvoiceDate ASC, inv.InvoiceID ASC, l.SrNo ASC
                """
            ),
            {"item_id": item_id, "date_from": date_from, "date_to": date_to},
        ).mappings().all()

        lines: list[dict[str, Any]] = []
        running = opening
        lines.append(
            self._decorate_line(
                {
                    "date": date_from.strftime("%d/%m/%Y"),
                    "description": "Opening Balance",
                    "debit": Decimal("0.00"),
                    "credit": Decimal("0.00"),
                    "balance": running,
                    "kind": "opening",
                }
            )
        )

        from flask import url_for

        for row in rows:
            credit = self._money(row["TaxableValue"])
            debit = Decimal("0.00")
            running = self._money(running + debit - credit)
            qty = self._money(row["Qty"])
            unit = (row["Unit"] or item["Unit"] or "").strip()
            cust = (row["CustomerName"] or "").strip()
            inv_no = (row["InvoiceNo"] or "").strip()
            particulars = (row["Particulars"] or item_name).strip()
            desc = f"{particulars} · Inv {inv_no}"
            if cust:
                desc = f"{desc} · {cust}"
            if qty:
                desc = f"{desc} · Qty {qty}" + (f" {unit}" if unit else "")
            txn_date = row["InvoiceDate"]
            invoice_id = int(row["InvoiceID"])
            lines.append(
                self._decorate_line(
                    {
                        "date": txn_date.strftime("%d/%m/%Y") if txn_date else "",
                        "description": desc,
                        "debit": debit,
                        "credit": credit,
                        "balance": running,
                        "kind": "txn",
                    },
                    link={
                        "can_open": True,
                        "source_module": "invoice",
                        "source_module_id": invoice_id,
                        "source_url": url_for(
                            "accounting_invoice.invoice_sale", edit=invoice_id
                        ),
                        "work_type": "",
                    },
                )
            )

        period = self._period_fields(date_from, date_to)
        return {
            "kind": "item",
            "title": "Item Ledger",
            "entity_name": item_name,
            "entity_id": item_id,
            "meta": [
                ("Item", item_name),
                ("Item Code", item_code or "—"),
                ("HSN / SAC", (item["HsnSac"] or "").strip() or "—"),
                ("Period", f"{date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}"),
            ],
            "headers": ["Date", "Description", "Debit", "Credit", "Closing Balance"],
            "lines": lines,
            "closing": running,
            **period,
        }

    # ── Export (PDF / XLSX / CSV / XML) ─────────────────────────────────────

    EXPORT_FORMATS = ("pdf", "xlsx", "csv", "xml")

    @staticmethod
    def _safe_filename(name: str) -> str:
        cleaned = re.sub(r"[^\w\-]+", "_", (name or "ledger").strip())[:50]
        return cleaned or "ledger"

    @staticmethod
    def _fmt_export_money(value: Any) -> str:
        if value is None:
            return ""
        return f"{Decimal(str(value or 0)):.2f}"

    def export_ledger(
        self,
        kind: str,
        entity_id: int,
        *,
        fmt: str,
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> tuple[bytes, str, str]:
        fmt_key = (fmt or "").strip().lower()
        if fmt_key not in self.EXPORT_FORMATS:
            raise ValueError("Unsupported export format. Use pdf, xlsx, csv, or xml.")
        ledger = self.preview_ledger(
            kind, entity_id, date_from=date_from, date_to=date_to
        )
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        base = (
            f"Ledger_{self._safe_filename(str(ledger.get('entity_name') or kind))}_{stamp}"
        )
        if fmt_key == "pdf":
            return self._export_pdf(ledger), f"{base}.pdf", "application/pdf"
        if fmt_key == "xlsx":
            return (
                self._export_xlsx(ledger),
                f"{base}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        if fmt_key == "csv":
            return self._export_csv(ledger), f"{base}.csv", "text/csv; charset=utf-8"
        return self._export_xml(ledger), f"{base}.xml", "application/xml"

    def _export_xlsx(self, ledger: dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger"
        headers = ledger.get("headers") or [
            "Date",
            "Description",
            "Debit",
            "Credit",
            "Closing Balance",
        ]
        thin = Border(
            left=Side(style="thin", color="5D6D7E"),
            right=Side(style="thin", color="5D6D7E"),
            top=Side(style="thin", color="5D6D7E"),
            bottom=Side(style="thin", color="5D6D7E"),
        )
        header_fill = PatternFill("solid", fgColor="154375")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        title_font = Font(name="Calibri", bold=True, size=14, color="154375")
        meta_font = Font(name="Calibri", size=10, color="1C2833")
        open_fill = PatternFill("solid", fgColor="EAF2F8")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = "Joshi Tax Consultancy & Services"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="148F77")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws["A2"] = f"{ledger.get('title') or 'Ledger'} — {ledger.get('entity_name') or ''}"
        ws["A2"].font = title_font

        row_idx = 4
        for label, value in ledger.get("meta") or []:
            ws.cell(row=row_idx, column=1, value=f"{label}:").font = Font(
                name="Calibri", bold=True, size=10
            )
            ws.merge_cells(
                start_row=row_idx, start_column=2, end_row=row_idx, end_column=len(headers)
            )
            cell = ws.cell(row=row_idx, column=2, value=str(value or ""))
            cell.font = meta_font
            row_idx += 1

        row_idx += 1
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
        row_idx += 1

        for line in ledger.get("lines") or []:
            values = [
                line.get("date") or "",
                line.get("description") or "",
                float(line["debit"]) if line.get("debit") is not None else None,
                float(line["credit"]) if line.get("credit") is not None else None,
                float(line["balance"]) if line.get("balance") is not None else None,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                if col >= 3 and value is not None:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                if (line.get("kind") or "") == "opening":
                    cell.fill = open_fill
                    cell.font = Font(name="Calibri", bold=True, size=10)
            row_idx += 1

        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Closing Balance").font = Font(
            name="Calibri", bold=True, size=11
        )
        close_cell = ws.cell(
            row=row_idx, column=5, value=float(self._money(ledger.get("closing")))
        )
        close_cell.font = Font(name="Calibri", bold=True, size=11, color="0E6655")
        close_cell.number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _export_csv(self, ledger: dict[str, Any]) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Joshi Tax Consultancy & Services"])
        writer.writerow([ledger.get("title") or "Ledger"])
        writer.writerow([ledger.get("entity_name") or ""])
        for label, value in ledger.get("meta") or []:
            writer.writerow([label, value])
        writer.writerow([])
        headers = ledger.get("headers") or [
            "Date",
            "Description",
            "Debit",
            "Credit",
            "Closing Balance",
        ]
        writer.writerow(headers)
        for line in ledger.get("lines") or []:
            writer.writerow(
                [
                    line.get("date") or "",
                    line.get("description") or "",
                    self._fmt_export_money(line.get("debit")),
                    self._fmt_export_money(line.get("credit")),
                    self._fmt_export_money(line.get("balance")),
                ]
            )
        writer.writerow([])
        writer.writerow(
            ["Closing Balance", "", "", "", self._fmt_export_money(ledger.get("closing"))]
        )
        return ("\ufeff" + buf.getvalue()).encode("utf-8")

    def _export_xml(self, ledger: dict[str, Any]) -> bytes:
        root = ET.Element("LedgerReport")
        ET.SubElement(root, "Company").text = "Joshi Tax Consultancy & Services"
        ET.SubElement(root, "Title").text = str(ledger.get("title") or "Ledger")
        ET.SubElement(root, "EntityName").text = str(ledger.get("entity_name") or "")
        ET.SubElement(root, "EntityId").text = str(ledger.get("entity_id") or "")
        ET.SubElement(root, "Kind").text = str(ledger.get("kind") or "")
        ET.SubElement(root, "ClosingBalance").text = self._fmt_export_money(
            ledger.get("closing")
        )

        meta_el = ET.SubElement(root, "Meta")
        for label, value in ledger.get("meta") or []:
            item = ET.SubElement(meta_el, "Item")
            item.set("label", str(label))
            item.text = str(value or "")

        lines_el = ET.SubElement(root, "Lines")
        for line in ledger.get("lines") or []:
            row = ET.SubElement(lines_el, "Line")
            row.set("kind", str(line.get("kind") or "txn"))
            ET.SubElement(row, "Date").text = str(line.get("date") or "")
            ET.SubElement(row, "Description").text = str(line.get("description") or "")
            ET.SubElement(row, "Debit").text = self._fmt_export_money(line.get("debit"))
            ET.SubElement(row, "Credit").text = self._fmt_export_money(line.get("credit"))
            ET.SubElement(row, "ClosingBalance").text = self._fmt_export_money(
                line.get("balance")
            )

        rough = ET.tostring(root, encoding="utf-8")
        pretty = minidom.parseString(rough).toprettyxml(indent="  ", encoding="utf-8")
        return pretty

    def _export_pdf(self, ledger: dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=14 * mm,
        )
        styles = getSampleStyleSheet()
        brand = ParagraphStyle(
            "LrBrand",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            textColor=colors.HexColor("#148F77"),
            spaceAfter=2,
        )
        title = ParagraphStyle(
            "LrTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=15,
            textColor=colors.HexColor("#154375"),
            spaceAfter=6,
        )
        meta_style = ParagraphStyle(
            "LrMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#1C2833"),
            leading=12,
        )
        cell_style = ParagraphStyle(
            "LrCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1C2833"),
        )

        story: list[Any] = [
            Paragraph("Joshi Tax Consultancy &amp; Services", brand),
            Paragraph(
                f"{ledger.get('title') or 'Ledger'} — {ledger.get('entity_name') or ''}",
                title,
            ),
        ]
        meta_bits = " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(
            f"<b>{label}:</b> {value}" for label, value in (ledger.get("meta") or [])
        )
        if meta_bits:
            story.append(Paragraph(meta_bits, meta_style))
        story.append(
            Paragraph(
                f"<b>Closing Balance:</b> Rs. {self._fmt_export_money(ledger.get('closing'))}",
                meta_style,
            )
        )
        story.append(Spacer(1, 8))

        table_data: list[list[Any]] = [
            ["Date", "Description", "Debit", "Credit", "Closing Balance"]
        ]
        for line in ledger.get("lines") or []:
            table_data.append(
                [
                    str(line.get("date") or ""),
                    Paragraph(
                        str(line.get("description") or "—")
                        .replace("&", "&amp;")
                        .replace("<", "&lt;")
                        .replace(">", "&gt;"),
                        cell_style,
                    ),
                    self._fmt_export_money(line.get("debit")),
                    self._fmt_export_money(line.get("credit")),
                    self._fmt_export_money(line.get("balance")),
                ]
            )

        table = Table(table_data, colWidths=[55, 250, 58, 58, 72])
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#154375")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#5D6D7E")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
        for i, line in enumerate(ledger.get("lines") or [], start=1):
            if (line.get("kind") or "") == "opening":
                style_cmds.append(
                    ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#EAF2F8"))
                )
                style_cmds.append(("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} · JTCS ERP",
                meta_style,
            )
        )
        doc.build(story)
        return buffer.getvalue()
