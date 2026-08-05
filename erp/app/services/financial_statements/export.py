"""PDF / Excel export for Financial Statements."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class FinancialStatementsExport:
    @staticmethod
    def _flatten_nodes(nodes: list[dict], depth: int = 0) -> list[tuple[str, str]]:
        rows: list[tuple[str, str]] = []
        for n in nodes or []:
            pad = "  " * depth
            rows.append((f"{pad}{n.get('group_name') or ''}", str(n.get("display_closing") or n.get("closing") or "0")))
            for led in n.get("ledgers") or []:
                rows.append(
                    (
                        f"{pad}  {led.get('ledger_name') or ''}",
                        str(led.get("display_closing") or led.get("closing") or "0"),
                    )
                )
            rows.extend(FinancialStatementsExport._flatten_nodes(n.get("children") or [], depth + 1))
        return rows

    @staticmethod
    def _split_pnl_sections(sections: list[dict]) -> tuple[list[dict], list[dict]]:
        left: list[dict] = []
        right: list[dict] = []
        for sec in sections or []:
            title = str(sec.get("title") or "")
            t = title.lower()
            try:
                amt = float(sec.get("total") if sec.get("total") is not None else sec.get("amount") or 0)
            except (TypeError, ValueError):
                amt = 0.0
            if sec.get("emphasis"):
                if amt >= 0:
                    right.append(sec)
                else:
                    left.append(
                        {
                            **sec,
                            "title": title.replace("Profit", "Loss").replace("profit", "loss"),
                            "total": abs(amt),
                        }
                    )
            elif "expense" in t:
                left.append(sec)
            else:
                right.append(sec)
        return left, right

    @classmethod
    def _sections_to_flat_rows(cls, sections: list[dict]) -> list[tuple[str, str]]:
        rows: list[tuple[str, str]] = []
        for sec in sections or []:
            rows.append((str(sec.get("title") or ""), str(sec.get("total") if sec.get("total") is not None else sec.get("amount") or "")))
            rows.extend(cls._flatten_nodes(sec.get("nodes") or []))
        return rows

    def to_excel(self, payload: dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        meta = payload.get("meta") or {}
        ws.title = (meta.get("report_title") or "Report")[:31]
        bold = Font(bold=True)
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws.append([meta.get("report_title") or "Financial Statement"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Period: {meta.get('date_from')} to {meta.get('date_to')}  |  {meta.get('fy_label')}"])
        ws.append([])

        layout = payload.get("layout")
        view = (payload.get("view") or "horizontal").strip().lower()
        if layout == "tally-bs":
            if view == "vertical":
                ws.append(["Particulars", "Amount"])
                for c in ws[4]:
                    c.font = bold
                ws.append(["LIABILITIES", ""])
                ws[ws.max_row][0].font = bold
                for name, amt in self._flatten_nodes(payload.get("left", {}).get("nodes") or []):
                    ws.append([name, amt])
                ws.append(["Total Liabilities", payload.get("left", {}).get("total")])
                ws.append([])
                ws.append(["ASSETS", ""])
                ws[ws.max_row][0].font = bold
                for name, amt in self._flatten_nodes(payload.get("right", {}).get("nodes") or []):
                    ws.append([name, amt])
                ws.append(["Total Assets", payload.get("right", {}).get("total")])
            else:
                ws.append(["Liabilities", "Amount", "Assets", "Amount"])
                for c in ws[4]:
                    c.font = bold
                left = self._flatten_nodes(payload.get("left", {}).get("nodes") or [])
                right = self._flatten_nodes(payload.get("right", {}).get("nodes") or [])
                for i in range(max(len(left), len(right))):
                    lname, lamt = left[i] if i < len(left) else ("", "")
                    rname, ramt = right[i] if i < len(right) else ("", "")
                    ws.append([lname, lamt, rname, ramt])
                ws.append(
                    [
                        "Total",
                        payload.get("left", {}).get("total"),
                        "Total",
                        payload.get("right", {}).get("total"),
                    ]
                )
        elif layout == "trial-balance":
            if view == "horizontal":
                ws.append(["Debit Particulars", "Debit", "Credit Particulars", "Credit"])
                for c in ws[4]:
                    c.font = bold
                dr = [
                    (r.get("ledger_name"), r.get("debit"))
                    for r in (payload.get("rows") or [])
                    if float(r.get("debit") or 0) > 0
                ]
                cr = [
                    (r.get("ledger_name"), r.get("credit"))
                    for r in (payload.get("rows") or [])
                    if float(r.get("credit") or 0) > 0
                ]
                for i in range(max(len(dr), len(cr), 1)):
                    dname, damt = dr[i] if i < len(dr) else ("", "")
                    cname, camt = cr[i] if i < len(cr) else ("", "")
                    ws.append([dname, damt, cname, camt])
                ws.append(
                    ["Total", payload.get("total_debit"), "Total", payload.get("total_credit")]
                )
            else:
                ws.append(["Ledger Name", "Group", "Debit", "Credit"])
                for c in ws[4]:
                    c.font = bold
                for r in payload.get("rows") or []:
                    ws.append(
                        [r.get("ledger_name"), r.get("group_name"), r.get("debit"), r.get("credit")]
                    )
                ws.append(["Total", "", payload.get("total_debit"), payload.get("total_credit")])
        elif layout in {"pnl", "trading"}:
            if view == "horizontal":
                left_secs, right_secs = self._split_pnl_sections(payload.get("sections") or [])
                ws.append(
                    [
                        "Expenses" if layout == "pnl" else "Debit",
                        "Amount",
                        "Income" if layout == "pnl" else "Credit",
                        "Amount",
                    ]
                )
                for c in ws[4]:
                    c.font = bold
                left_rows = self._sections_to_flat_rows(left_secs)
                right_rows = self._sections_to_flat_rows(right_secs)
                for i in range(max(len(left_rows), len(right_rows), 1)):
                    lname, lamt = left_rows[i] if i < len(left_rows) else ("", "")
                    rname, ramt = right_rows[i] if i < len(right_rows) else ("", "")
                    ws.append([lname, lamt, rname, ramt])
            else:
                for sec in payload.get("sections") or []:
                    ws.append([sec.get("title"), sec.get("total")])
                    ws[ws.max_row][0].font = bold
                    for name, amt in self._flatten_nodes(sec.get("nodes") or []):
                        ws.append([name, amt])
                    ws.append([])
        elif layout in {"depreciation", "fixed-assets"}:
            ws.append(
                [
                    "Asset",
                    "Purchase Date",
                    "Purchase Value",
                    "Rate %",
                    "CY Depreciation",
                    "Accumulated",
                    "WDV",
                ]
            )
            for c in ws[4]:
                c.font = bold
            for r in payload.get("rows") or []:
                ws.append(
                    [
                        r.get("asset_name"),
                        r.get("purchase_date"),
                        r.get("purchase_value"),
                        r.get("depreciation_rate"),
                        r.get("current_year_depreciation"),
                        r.get("accumulated_depreciation"),
                        r.get("wdv"),
                    ]
                )
        elif layout == "ratios":
            ws.append(["Ratio", "Value", "Formula"])
            for c in ws[4]:
                c.font = bold
            for r in payload.get("ratios") or []:
                ws.append([r.get("name"), r.get("value"), r.get("formula")])
        else:
            ws.append(["Section", "Amount"])
            for sec in payload.get("sections") or []:
                ws.append([sec.get("title"), sec.get("amount") or sec.get("total")])
            for r in payload.get("sources") or []:
                ws.append([f"Source: {r.get('name')}", r.get("amount")])
            for r in payload.get("applications") or []:
                ws.append([f"Application: {r.get('name')}", r.get("amount")])

        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col[:50]) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 48)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def to_pdf(self, payload: dict[str, Any]) -> bytes:
        meta = payload.get("meta") or {}
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("FSTitle", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
        story = [
            Paragraph(meta.get("report_title") or "Financial Statement", title_style),
            Paragraph(
                f"Period: {meta.get('date_from')} to {meta.get('date_to')} &nbsp;|&nbsp; {meta.get('fy_label')}",
                styles["Normal"],
            ),
            Spacer(1, 8),
        ]

        layout = payload.get("layout")
        view = (payload.get("view") or "horizontal").strip().lower()
        data: list[list[Any]] = []
        if layout == "tally-bs":
            if view == "vertical":
                data.append(["Particulars", "Amount"])
                data.append(["LIABILITIES", ""])
                for name, amt in self._flatten_nodes(payload.get("left", {}).get("nodes") or []):
                    data.append([name, amt])
                data.append(["Total Liabilities", payload.get("left", {}).get("total")])
                data.append(["ASSETS", ""])
                for name, amt in self._flatten_nodes(payload.get("right", {}).get("nodes") or []):
                    data.append([name, amt])
                data.append(["Total Assets", payload.get("right", {}).get("total")])
            else:
                data.append(["Liabilities", "Amount", "Assets", "Amount"])
                left = self._flatten_nodes(payload.get("left", {}).get("nodes") or [])
                right = self._flatten_nodes(payload.get("right", {}).get("nodes") or [])
                for i in range(max(len(left), len(right), 1)):
                    lname, lamt = left[i] if i < len(left) else ("", "")
                    rname, ramt = right[i] if i < len(right) else ("", "")
                    data.append([lname, lamt, rname, ramt])
                data.append(
                    [
                        "Total",
                        payload.get("left", {}).get("total"),
                        "Total",
                        payload.get("right", {}).get("total"),
                    ]
                )
        elif layout == "trial-balance":
            if view == "horizontal":
                data.append(["Debit Particulars", "Debit", "Credit Particulars", "Credit"])
                dr = [
                    (r.get("ledger_name"), r.get("debit"))
                    for r in (payload.get("rows") or [])
                    if float(r.get("debit") or 0) > 0
                ]
                cr = [
                    (r.get("ledger_name"), r.get("credit"))
                    for r in (payload.get("rows") or [])
                    if float(r.get("credit") or 0) > 0
                ]
                for i in range(max(len(dr), len(cr), 1)):
                    dname, damt = dr[i] if i < len(dr) else ("", "")
                    cname, camt = cr[i] if i < len(cr) else ("", "")
                    data.append([dname, damt, cname, camt])
                data.append(
                    ["Total", payload.get("total_debit"), "Total", payload.get("total_credit")]
                )
            else:
                data.append(["Ledger Name", "Group", "Debit", "Credit"])
                for r in payload.get("rows") or []:
                    data.append(
                        [r.get("ledger_name"), r.get("group_name"), r.get("debit"), r.get("credit")]
                    )
                data.append(["Total", "", payload.get("total_debit"), payload.get("total_credit")])
        elif layout in {"pnl", "trading"} and view == "horizontal":
            left_secs, right_secs = self._split_pnl_sections(payload.get("sections") or [])
            data.append(
                [
                    "Expenses" if layout == "pnl" else "Debit",
                    "Amount",
                    "Income" if layout == "pnl" else "Credit",
                    "Amount",
                ]
            )
            left_rows = self._sections_to_flat_rows(left_secs)
            right_rows = self._sections_to_flat_rows(right_secs)
            for i in range(max(len(left_rows), len(right_rows), 1)):
                lname, lamt = left_rows[i] if i < len(left_rows) else ("", "")
                rname, ramt = right_rows[i] if i < len(right_rows) else ("", "")
                data.append([lname, lamt, rname, ramt])
        else:
            data.append(["Particulars", "Amount"])
            for sec in payload.get("sections") or []:
                data.append([sec.get("title"), sec.get("amount") or sec.get("total")])
                for name, amt in self._flatten_nodes(sec.get("nodes") or []):
                    data.append([name, amt])
            for r in payload.get("rows") or []:
                data.append([r.get("asset_name") or r.get("ledger_name") or r.get("name"), r.get("wdv") or r.get("value") or r.get("debit")])

        if data:
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        return buf.getvalue()
