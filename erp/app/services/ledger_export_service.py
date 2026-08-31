"""Admin Role → Import/Export → Ledger Export (styled Excel + watermarked PDF)."""

from __future__ import annotations

import io
import math
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import text

from app.extensions import db
from app.utils.opening_balance import apply_account_running, is_credit_normal_nature
from app.services.payment_accounting_service import (
    sql_customer_receipt_expr,
    sql_unpaid_followup_exclusion,
)

# Brand palette (professional, colourful — not purple/glow AI defaults)
COLOR_NAVY = "1B4F72"
COLOR_TEAL = "148F77"
COLOR_HEADER_BG = "1A5276"
COLOR_HEADER_FG = "FFFFFF"
COLOR_META_LABEL = "1B4F72"
COLOR_META_BG = "EBF5FB"
COLOR_ALT_ROW = "E8F6F3"
COLOR_OPENING = "FCF3CF"
COLOR_CLOSING = "D5F5E3"
COLOR_DEBIT = "922B21"
COLOR_CREDIT = "196F3D"
COLOR_BALANCE = "1A5276"
COLOR_BORDER = "5D6D7E"


class LedgerExportService:
    @staticmethod
    def _money(value) -> Decimal:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))

    @staticmethod
    def _parse_date(raw: str | None, fallback: date) -> date:
        value = (raw or "").strip()
        if not value:
            return fallback
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return fallback

    @staticmethod
    def _last4(masked: str | None, account_number: str | None) -> str:
        from app.services.dashboard_service import DashboardService

        return DashboardService._account_suffix(account_number, masked)

    @staticmethod
    def _fmt_money(value: Decimal | float | int) -> str:
        return f"{Decimal(str(value or 0)):.2f}"

    @staticmethod
    def _is_credit_normal_nature(nature: str | None, under_type: str | None = None) -> bool:
        return is_credit_normal_nature(nature, under_type)

    def _apply_bank_running(
        self,
        previous: Decimal,
        debit: Decimal,
        credit: Decimal,
        *,
        credit_normal: bool,
    ) -> Decimal:
        return apply_account_running(
            previous, debit, credit, credit_normal=credit_normal
        )

    def list_bank_accounts(self, *, search: str | None = None) -> list[dict[str, Any]]:
        params: dict[str, Any] = {}
        search_sql = ""
        needle = (search or "").strip()
        if needle:
            params["like"] = f"%{needle}%"
            search_sql = """
              AND (
                a.BankName LIKE :like
                OR a.AccountNumber LIKE :like
                OR a.MaskedAccountNumber LIKE :like
                OR a.AccountHolderName LIKE :like
                OR a.AccountType LIKE :like
              )
            """
        rows = db.session.execute(
            text(
                f"""
                SELECT
                    a.JtcsBankAccountID AS account_id,
                    a.BankName,
                    a.MaskedAccountNumber,
                    a.AccountNumber,
                    a.AccountType,
                    a.AccountHolderName,
                    ISNULL(a.OpeningBalance, 0) AS OpeningBalance,
                    a.OpeningBalanceDate,
                    ISNULL(a.ActiveStatus, 1) AS ActiveStatus,
                    (
                        SELECT COUNT(1)
                        FROM dbo.JtcsBankTransaction t
                        WHERE t.JtcsBankAccountID = a.JtcsBankAccountID
                    ) AS txn_count
                FROM dbo.JtcsBankAccountMaster a
                WHERE 1 = 1
                  {search_sql}
                ORDER BY
                    CASE WHEN LOWER(LTRIM(RTRIM(ISNULL(a.BankName, N'')))) = N'cash' THEN 0 ELSE 1 END,
                    a.DisplayOrder,
                    a.BankName,
                    a.JtcsBankAccountID
                """
            ),
            params,
        ).mappings().all()

        result = []
        for row in rows:
            bank_name = (row["BankName"] or "Account").strip() or "Account"
            last4 = self._last4(row["MaskedAccountNumber"], row["AccountNumber"])
            account_type = (row["AccountType"] or "").strip()
            label_parts = [bank_name]
            if last4:
                label_parts.append(f"({last4})")
            if account_type:
                label_parts.append(f"[{account_type}]")
            result.append(
                {
                    "account_id": int(row["account_id"]),
                    "bank_name": bank_name,
                    "masked_account": (row["MaskedAccountNumber"] or row["AccountNumber"] or "").strip(),
                    "account_type": account_type,
                    "account_holder": (row["AccountHolderName"] or "").strip(),
                    "label": " ".join(label_parts),
                    "opening_balance": str(self._money(row["OpeningBalance"])),
                    "txn_count": int(row["txn_count"] or 0),
                    "active": bool(row["ActiveStatus"]),
                }
            )
        return result

    def list_customers(self, *, search: str | None = None, limit: int = 200) -> list[dict[str, Any]]:
        params: dict[str, Any] = {"lim": max(1, min(int(limit or 200), 500))}
        search_sql = ""
        needle = (search or "").strip()
        if needle:
            mobile = re.sub(r"\D", "", needle)
            params["like"] = f"%{needle}%"
            params["like_upper"] = f"%{needle.upper()}%"
            params["mobile_like"] = f"%{mobile}%" if mobile else f"%{needle}%"
            search_sql = """
              AND (
                UPPER(LTRIM(RTRIM(c.CustomerName))) LIKE UPPER(:like)
                OR ISNULL(c.MobileNumber, N'') LIKE :mobile_like
                OR UPPER(LTRIM(RTRIM(ISNULL(c.PANNumber, N'')))) LIKE :like_upper
              )
            """

        # Daily txns + unpaid followup bills + Bank/Cash electronic transfers on customer CoA.
        obc_count_sql = "0"
        followup_count_sql = "0"
        try:
            has_obc_keys = bool(
                db.session.execute(
                    text(
                        "SELECT CASE WHEN COL_LENGTH(N'dbo.OthersBankCashTransaction', "
                        "N'CreditLedgerKey') IS NULL THEN 0 ELSE 1 END"
                    )
                ).scalar()
            )
            has_coa_customer = bool(
                db.session.execute(
                    text(
                        "SELECT CASE WHEN COL_LENGTH(N'dbo.ChartOfAccountMaster', "
                        "N'CustomerID') IS NULL THEN 0 ELSE 1 END"
                    )
                ).scalar()
            )
            if has_obc_keys and has_coa_customer:
                obc_count_sql = """
                    (
                        SELECT COUNT(1)
                        FROM dbo.OthersBankCashTransaction e
                        INNER JOIN dbo.ChartOfAccountMaster a
                            ON a.CustomerID = c.CustomerID
                           AND ISNULL(a.IsActive, 1) = 1
                        WHERE ISNULL(e.IsActive, 1) = 1
                          AND (
                                e.CreditLedgerKey = CONCAT(N'coa-', a.AccountID)
                             OR e.DebitLedgerKey = CONCAT(N'coa-', a.AccountID)
                          )
                    )
                """
            if db.session.execute(
                text("SELECT OBJECT_ID(N'dbo.FollowupEntryMaster', N'U')")
            ).scalar():
                followup_count_sql = """
                    (
                        SELECT COUNT(1)
                        FROM dbo.FollowupEntryMaster f
                        WHERE f.CustomerID = c.CustomerID
                          AND ISNULL(f.IsActive, 1) = 1
                          AND f.BillNo IS NOT NULL
                          AND LTRIM(RTRIM(f.BillNo)) <> N''
                          AND ISNULL(f.BillAmount, 0) > 0
                          AND NOT EXISTS (
                              SELECT 1
                              FROM dbo.JTCSDailyTransaction d2
                              WHERE d2.CustomerID = f.CustomerID
                                AND d2.Status = N'Posted'
                                AND UPPER(LTRIM(RTRIM(ISNULL(d2.ReferenceNo, N''))))
                                    = UPPER(LTRIM(RTRIM(f.BillNo)))
                                AND d2.WorkType = f.ModuleCode
                          )
                    )
                """
        except Exception:
            db.session.rollback()
            obc_count_sql = "0"
            followup_count_sql = "0"

        rows = db.session.execute(
            text(
                f"""
                SELECT TOP (:lim)
                    c.CustomerID,
                    c.CustomerName,
                    c.MobileNumber,
                    c.PANNumber,
                    c.CustomerStatus,
                    (
                        SELECT COUNT(1)
                        FROM dbo.JTCSDailyTransaction d
                        WHERE d.CustomerID = c.CustomerID
                          AND d.Status = N'Posted'
                    )
                    + {obc_count_sql}
                    + {followup_count_sql}
                    AS txn_count
                FROM dbo.CustomerMaster c
                WHERE ISNULL(c.CustomerStatus, N'Active') <> N'Rejected'
                  {search_sql}
                ORDER BY
                    CASE WHEN ISNULL(c.CustomerStatus, N'Active') = N'Active' THEN 0 ELSE 1 END,
                    c.CustomerName,
                    c.CustomerID
                """
            ),
            params,
        ).mappings().all()
        return [
            {
                "customer_id": int(row["CustomerID"]),
                "customer_name": (row["CustomerName"] or "").strip(),
                "mobile_number": (row["MobileNumber"] or "").strip(),
                "pan_number": (row["PANNumber"] or "").strip(),
                "status": (row["CustomerStatus"] or "").strip(),
                "txn_count": int(row["txn_count"] or 0),
            }
            for row in rows
        ]

    def _resolve_period(
        self, date_from: date | None, date_to: date | None
    ) -> tuple[date, date]:
        today = date.today()
        return date_from or date(2000, 1, 1), date_to or today

    def _load_bank_account_for_ledger(self, account_id: int):
        """Bank Master row plus Chart of Group nature (Asset/Liability/Income/Expense)."""
        has_group = db.session.execute(
            text(
                """
                SELECT CASE
                    WHEN COL_LENGTH(N'dbo.JtcsBankAccountMaster', N'ChartGroupID') IS NULL THEN 0
                    WHEN OBJECT_ID(N'dbo.ChartOfGroupMaster', N'U') IS NULL THEN 0
                    ELSE 1
                END
                """
            )
        ).scalar()
        if has_group:
            return db.session.execute(
                text(
                    """
                    SELECT
                        a.JtcsBankAccountID, a.BankName, a.MaskedAccountNumber, a.AccountNumber,
                        a.AccountType, a.AccountHolderName, a.OpeningBalance, a.OpeningBalanceDate,
                        a.ChartGroupID,
                        g.GroupName,
                        g.UnderType,
                        ISNULL(
                            NULLIF(g.GroupNature, N''),
                            CASE
                                WHEN g.UnderType = N'Liabilities' THEN N'Liability'
                                WHEN g.UnderType = N'Assets' THEN N'Asset'
                                ELSE N'Asset'
                            END
                        ) AS GroupNature
                    FROM dbo.JtcsBankAccountMaster a
                    LEFT JOIN dbo.ChartOfGroupMaster g ON g.GroupID = a.ChartGroupID
                    WHERE a.JtcsBankAccountID = :account_id
                    """
                ),
                {"account_id": account_id},
            ).mappings().first()
        return db.session.execute(
            text(
                """
                SELECT JtcsBankAccountID, BankName, MaskedAccountNumber, AccountNumber,
                       AccountType, AccountHolderName, OpeningBalance, OpeningBalanceDate,
                       CAST(NULL AS INT) AS ChartGroupID,
                       CAST(NULL AS NVARCHAR(150)) AS GroupName,
                       CAST(NULL AS NVARCHAR(20)) AS UnderType,
                       CAST(N'Asset' AS NVARCHAR(20)) AS GroupNature
                FROM dbo.JtcsBankAccountMaster
                WHERE JtcsBankAccountID = :account_id
                """
            ),
            {"account_id": account_id},
        ).mappings().first()

    @staticmethod
    def _to_ledger_date(value) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text_val = str(value).strip()
        if not text_val:
            return None
        try:
            return date.fromisoformat(text_val[:10])
        except ValueError:
            return None

    @staticmethod
    def _alias_snapshot_sql(column_prefix: str = "") -> str:
        col_mask = f"{column_prefix}MaskedAccountNumber" if column_prefix else "MaskedAccountNumber"
        col_name = f"{column_prefix}BankName" if column_prefix else "BankName"
        return f"""
            (
                REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL({col_mask}, N'')))), N' ', N''), N'-', N'')
                    LIKE :wallet_alias_like
                OR REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL({col_name}, N'')))), N' ', N''), N'-', N'')
                    LIKE :wallet_alias_like
            )
        """

    def _wallet_ledger_flags(self, account_id: int) -> dict[str, Any]:
        from app.utils.shcil_bank_accounts import (
            account_is_ecourt_purchase_wallet,
            account_is_stamp_purchase_wallet,
        )

        is_stamp = False
        is_ecourt = False
        try:
            is_stamp = account_is_stamp_purchase_wallet(db.session, account_id)
        except Exception:
            is_stamp = False
        try:
            is_ecourt = account_is_ecourt_purchase_wallet(db.session, account_id)
        except Exception:
            is_ecourt = False
        return {"stamp": is_stamp, "ecourt": is_ecourt}

    def _bank_ledger_account_where(
        self, account_id: int, wallet: dict[str, Any]
    ) -> tuple[str, dict[str, Any]]:
        """Match this bank account only.

        Stamp / e-Court wallets also pick up snapshot rows stored with their
        live account numbers (0213UK1423304 / HUKECFUK1423304). Other banks'
        transactions are never rewritten or reassigned.
        """
        from app.utils.shcil_bank_accounts import (
            ECOURT_PURCHASE_ACCOUNT_NUMBERS,
            STAMP_PURCHASE_ACCOUNT_NUMBERS,
        )

        params: dict[str, Any] = {"account_id": account_id}
        clauses = ["JtcsBankAccountID = :account_id"]
        account_numbers: tuple[str, ...] = ()
        if wallet.get("stamp"):
            account_numbers = STAMP_PURCHASE_ACCOUNT_NUMBERS
        elif wallet.get("ecourt"):
            account_numbers = ECOURT_PURCHASE_ACCOUNT_NUMBERS
        if account_numbers:
            number = "".join(ch for ch in account_numbers[0].upper() if ch.isalnum())
            clauses.append(
                """
                (
                    REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL(MaskedAccountNumber, N'')))), N' ', N''), N'-', N'')
                        = :wallet_account_number
                    OR REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL(BankName, N'')))), N' ', N''), N'-', N'')
                        = :wallet_account_number
                )
                """
            )
            params["wallet_account_number"] = number
        where_sql = "(" + " OR ".join(clauses) + ")"
        return where_sql, params

    def _missing_purpose_purchase_sql(self) -> str:
        return """
            SELECT
                d.TransactionID,
                d.TransactionDate,
                CASE
                    WHEN ISNULL(d.PurchaseAmount, 0) > 0 THEN ISNULL(d.PurchaseAmount, 0)
                    ELSE ISNULL(s.StampDutyAmount, 0)
                END AS PurchaseAmount,
                d.ReferenceNo,
                d.Remarks
            FROM dbo.JTCSDailyTransaction d
            LEFT JOIN dbo.StampMaster s
                ON s.StampID = d.StampID
            WHERE d.WorkType = N'SHCIL'
              AND d.SubWorkType = :sub_work
              AND (
                    ISNULL(d.PurchaseAmount, 0) > 0
                    OR ISNULL(s.StampDutyAmount, 0) > 0
                  )
              AND d.TransactionDate >= :range_from
              AND d.TransactionDate < :range_to
              AND NOT EXISTS (
                    SELECT 1
                    FROM dbo.JtcsBankTransaction t
                    WHERE (
                            t.SourceRecordID = d.TransactionID
                         OR t.SourceID = d.TransactionID
                      )
                      AND t.JtcsBankAccountID = :account_id
                      AND UPPER(LTRIM(RTRIM(ISNULL(t.Description, N'')))) = :purchase_desc
                      AND ISNULL(t.Credit, 0) > 0
              )
            ORDER BY d.TransactionDate ASC, d.TransactionID ASC
        """

    def _ensure_purpose_purchase_legs(self, account_id: int, wallet: dict[str, Any]) -> None:
        """Post missing stamp-duty / e-Court purchase credits onto this wallet only.

        Inserts new JtcsBankTransaction rows for this account_id. Does not
        update, delete, or reassign any other bank account's transactions.
        """
        if not wallet.get("stamp") and not wallet.get("ecourt"):
            return

        account_row = db.session.execute(
            text(
                """
                SELECT BankName, AccountNumber, MaskedAccountNumber
                FROM dbo.JtcsBankAccountMaster
                WHERE JtcsBankAccountID = :account_id
                """
            ),
            {"account_id": account_id},
        ).mappings().first()
        if account_row is None:
            return

        if wallet.get("stamp"):
            from app.utils.shcil_bank_accounts import STAMP_PURCHASE_DESCRIPTION

            sub_work = "Stamp Activity"
            description = "Stamp Purchase"
            purchase_desc = STAMP_PURCHASE_DESCRIPTION
        else:
            from app.utils.shcil_bank_accounts import ECOURT_PURCHASE_DESCRIPTION

            sub_work = "e-Court Activity"
            description = "e-Court Purchase"
            purchase_desc = ECOURT_PURCHASE_DESCRIPTION

        bank_name = (account_row["BankName"] or "").strip() or "Bank"
        masked = (
            (account_row["AccountNumber"] or "").strip()
            or (account_row["MaskedAccountNumber"] or "").strip()
            or "NA"
        )
        payment_mode = db.session.execute(
            text(
                """
                SELECT TOP 1 PaymentModeID
                FROM dbo.PaymentModeMaster
                WHERE BankAccountID = :account_id
                  AND ISNULL(IsActive, 1) = 1
                ORDER BY PaymentModeID
                """
            ),
            {"account_id": account_id},
        ).first()
        payment_mode_id = int(payment_mode[0]) if payment_mode and payment_mode[0] else None

        try:
            result = db.session.execute(
                text(
                    """
                    INSERT INTO dbo.JtcsBankTransaction (
                        JtcsBankAccountID,
                        BankName,
                        MaskedAccountNumber,
                        TransactionDate,
                        Description,
                        Debit,
                        Credit,
                        ClosingBalance,
                        ImportedBy,
                        ImportedDate,
                        Remarks,
                        IsLocked,
                        SourceTable,
                        SourceRecordID,
                        SourceType,
                        SourceID,
                        LedgerKind,
                        PaymentModeID,
                        PaymentSequence
                    )
                    SELECT
                        :account_id,
                        :bank_name,
                        :masked,
                        d.TransactionDate,
                        :description,
                        NULL,
                        CASE
                            WHEN ISNULL(d.PurchaseAmount, 0) > 0 THEN ISNULL(d.PurchaseAmount, 0)
                            ELSE ISNULL(s.StampDutyAmount, 0)
                        END,
                        0,
                        N'System',
                        GETUTCDATE(),
                        d.ReferenceNo,
                        0,
                        N'JTCSDailyTransaction',
                        d.TransactionID,
                        N'SHCIL',
                        d.TransactionID,
                        N'PAYMENT',
                        :payment_mode_id,
                        ISNULL(d.PaymentSplitCount, 1) + 1
                    FROM dbo.JTCSDailyTransaction d
                    LEFT JOIN dbo.StampMaster s
                        ON s.StampID = d.StampID
                    WHERE d.WorkType = N'SHCIL'
                      AND d.SubWorkType = :sub_work
                      AND (
                            ISNULL(d.PurchaseAmount, 0) > 0
                            OR ISNULL(s.StampDutyAmount, 0) > 0
                          )
                      AND NOT EXISTS (
                            SELECT 1
                            FROM dbo.JtcsBankTransaction t
                            WHERE (
                                    t.SourceRecordID = d.TransactionID
                                 OR t.SourceID = d.TransactionID
                              )
                              AND t.JtcsBankAccountID = :account_id
                              AND UPPER(LTRIM(RTRIM(ISNULL(t.Description, N'')))) = :purchase_desc
                              AND ISNULL(t.Credit, 0) > 0
                      )
                    """
                ),
                {
                    "account_id": account_id,
                    "bank_name": bank_name[:150],
                    "masked": masked[:50],
                    "description": description,
                    "payment_mode_id": payment_mode_id,
                    "sub_work": sub_work,
                    "purchase_desc": purchase_desc,
                },
            )
            if result.rowcount:
                db.session.commit()
        except Exception:
            db.session.rollback()

    def _append_missing_purpose_purchases(
        self,
        rows: list[dict[str, Any]],
        *,
        account_id: int,
        sub_work: str,
        purchase_description: str,
        description: str,
        range_from: date,
        range_to_next: date,
        ob_date: date | None,
        seen_ids: set[int],
    ) -> tuple[Decimal, Decimal]:
        extra_debit = Decimal("0.00")
        extra_credit = Decimal("0.00")
        fetched = db.session.execute(
            text(self._missing_purpose_purchase_sql()),
            {
                "account_id": account_id,
                "sub_work": sub_work,
                "purchase_desc": purchase_description,
                "range_from": range_from,
                "range_to": range_to_next,
            },
        ).mappings().all()
        for raw in fetched:
            txn_date = self._to_ledger_date(raw["TransactionDate"])
            if txn_date is None:
                continue
            if ob_date is not None and txn_date < ob_date:
                continue
            amount = self._money(raw["PurchaseAmount"])
            if amount <= 0:
                continue
            daily_id = int(raw["TransactionID"] or 0)
            if daily_id in seen_ids:
                continue
            seen_ids.add(daily_id)
            extra_credit = self._money(extra_credit + amount)
            rows.append(
                {
                    "JtcsBankTransactionID": -daily_id,
                    "TransactionDate": txn_date,
                    "Description": description,
                    "Remarks": (raw["ReferenceNo"] or raw["Remarks"] or "").strip() or None,
                    "SourceTable": "JTCSDailyTransaction",
                    "SourceType": "SHCIL",
                    "SourceRecordID": daily_id,
                    "LedgerKind": "PAYMENT",
                    "DebitValue": Decimal("0.00"),
                    "CreditValue": amount,
                }
            )
        return extra_debit, extra_credit

    def _append_orphan_obc_deposits(
        self,
        rows: list[dict[str, Any]],
        *,
        account_id: int,
        range_from: date,
        range_to_next: date,
        ob_date: date | None,
    ) -> tuple[Decimal, Decimal]:
        extra_debit = Decimal("0.00")
        extra_credit = Decimal("0.00")
        fetched = db.session.execute(
            text(
                """
                SELECT
                    o.EntryID,
                    o.VoucherNo,
                    o.WorkDate,
                    o.Amount,
                    o.Purpose,
                    o.Remarks
                FROM dbo.OthersBankCashTransaction o
                WHERE o.DebitBankAccountID = :account_id
                  AND ISNULL(o.IsActive, 1) = 1
                  AND o.WorkDate >= :range_from
                  AND o.WorkDate < :range_to
                  AND (
                        o.InBankTransactionID IS NULL
                        OR NOT EXISTS (
                            SELECT 1
                            FROM dbo.JtcsBankTransaction t2
                            WHERE t2.JtcsBankTransactionID = o.InBankTransactionID
                              AND t2.JtcsBankAccountID = :account_id
                        )
                  )
                ORDER BY o.WorkDate ASC, o.EntryID ASC
                """
            ),
            {
                "account_id": account_id,
                "range_from": range_from,
                "range_to": range_to_next,
            },
        ).mappings().all()
        for raw in fetched:
            txn_date = self._to_ledger_date(raw["WorkDate"])
            if txn_date is None:
                continue
            if ob_date is not None and txn_date < ob_date:
                continue
            amount = self._money(raw["Amount"])
            if amount <= 0:
                continue
            extra_debit = self._money(extra_debit + amount)
            purpose = (raw["Purpose"] or "Bank Transfer").strip()
            voucher = (raw["VoucherNo"] or "").strip()
            rows.append(
                {
                    "JtcsBankTransactionID": -int(raw["EntryID"] or 0),
                    "TransactionDate": txn_date,
                    "Description": f"{purpose} (Debit / In) — OBC deposit",
                    "Remarks": voucher or (raw["Remarks"] or None),
                    "SourceTable": "OthersBankCashTransaction",
                    "SourceType": "OTHERS_BANK_CASH",
                    "SourceRecordID": int(raw["EntryID"] or 0),
                    "LedgerKind": "CONTRA_IN",
                    "DebitValue": amount,
                    "CreditValue": Decimal("0.00"),
                }
            )
        return extra_debit, extra_credit

    def _dedupe_bank_txn_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen: set[int] = set()
        unique: list[dict[str, Any]] = []
        for row in rows:
            tid = int(row.get("JtcsBankTransactionID") or 0)
            if tid > 0:
                if tid in seen:
                    continue
                seen.add(tid)
            unique.append(row)
        unique.sort(
            key=lambda r: (
                self._to_ledger_date(r.get("TransactionDate")) or date.min,
                int(r.get("JtcsBankTransactionID") or 0),
            )
        )
        return unique

    def _bank_ledger_data(
        self,
        account_id: int,
        *,
        date_from: date | None,
        date_to: date | None,
    ) -> dict[str, Any]:
        account = self._load_bank_account_for_ledger(account_id)
        if account is None:
            raise ValueError("Bank account not found.")

        credit_normal = self._is_credit_normal_nature(
            account.get("GroupNature"), account.get("UnderType")
        )

        date_from, date_to = self._resolve_period(date_from, date_to)
        bank_name = (account["BankName"] or "Account").strip() or "Account"
        last4 = self._last4(account["MaskedAccountNumber"], account["AccountNumber"])
        account_type = (account["AccountType"] or "").strip()
        label_parts = [bank_name]
        if last4:
            label_parts.append(f"({last4})")
        if account_type:
            label_parts.append(f"[{account_type}]")
        label = " ".join(label_parts)

        opening = Decimal("0.00")
        ob_date = self._to_ledger_date(account["OpeningBalanceDate"])
        if ob_date is None or ob_date <= date_from:
            opening = self._money(account["OpeningBalance"])
        wallet = self._wallet_ledger_flags(account_id)
        self._ensure_purpose_purchase_legs(account_id, wallet)
        account_where, account_params = self._bank_ledger_account_where(account_id, wallet)
        date_to_next = date_to + timedelta(days=1)
        # Inclusive To Date: datetime rows on date_to itself used to be dropped
        # by TransactionDate <= midnight. Use half-open [from, to+1).
        # Prior movements only on/after OpeningBalanceDate — never double-count
        # Bank Master Opening Balance with pre-opening (or corrupt-dated) rows.
        prior_sql = f"""
                SELECT
                    ISNULL(SUM(ISNULL(Debit, 0)), 0) AS prior_debit,
                    ISNULL(SUM(ISNULL(Credit, 0)), 0) AS prior_credit
                FROM dbo.JtcsBankTransaction
                WHERE {account_where}
                  AND TransactionDate >= :range_from
                  AND TransactionDate < :range_to
            """
        prior_from = ob_date or date(2000, 1, 1)
        prior_params = {
            **account_params,
            "range_from": prior_from,
            "range_to": date_from,
        }
        prior_row = db.session.execute(text(prior_sql), prior_params).mappings().first()
        prior_debit = self._money(prior_row["prior_debit"] if prior_row else 0)
        prior_credit = self._money(prior_row["prior_credit"] if prior_row else 0)

        txn_sql = f"""
                SELECT
                    JtcsBankTransactionID, TransactionDate, Description, Remarks,
                    SourceTable, SourceType, SourceRecordID, LedgerKind,
                    ISNULL(Debit, 0) AS DebitValue,
                    ISNULL(Credit, 0) AS CreditValue
                FROM dbo.JtcsBankTransaction
                WHERE {account_where}
                  AND TransactionDate >= :range_from
                  AND TransactionDate < :range_to
                ORDER BY TransactionDate ASC, JtcsBankTransactionID ASC
            """
        txn_params = {
            **account_params,
            "range_from": date_from if ob_date is None else max(date_from, ob_date),
            "range_to": date_to_next,
        }
        txn_rows = [
            {
                **dict(raw),
                "TransactionDate": self._to_ledger_date(raw["TransactionDate"]),
            }
            for raw in db.session.execute(text(txn_sql), txn_params).mappings().all()
        ]
        seen_purchase_dailies = {
            int(row["SourceRecordID"])
            for row in txn_rows
            if row.get("SourceRecordID")
            and (row.get("Description") or "").strip().upper()
            in {"STAMP PURCHASE", "E-COURT PURCHASE"}
        }
        extra_prior_debit = Decimal("0.00")
        extra_prior_credit = Decimal("0.00")
        if wallet.get("stamp"):
            from app.utils.shcil_bank_accounts import STAMP_PURCHASE_DESCRIPTION

            extra_prior_debit, extra_prior_credit = self._append_missing_purpose_purchases(
                [],
                account_id=account_id,
                sub_work="Stamp Activity",
                purchase_description=STAMP_PURCHASE_DESCRIPTION,
                description="Stamp Purchase",
                range_from=prior_from,
                range_to_next=date_from,
                ob_date=ob_date,
                seen_ids=seen_purchase_dailies,
            )
            self._append_missing_purpose_purchases(
                txn_rows,
                account_id=account_id,
                sub_work="Stamp Activity",
                purchase_description=STAMP_PURCHASE_DESCRIPTION,
                description="Stamp Purchase",
                range_from=txn_params["range_from"],
                range_to_next=date_to_next,
                ob_date=ob_date,
                seen_ids=seen_purchase_dailies,
            )
        elif wallet.get("ecourt"):
            from app.utils.shcil_bank_accounts import ECOURT_PURCHASE_DESCRIPTION

            extra_prior_debit, extra_prior_credit = self._append_missing_purpose_purchases(
                [],
                account_id=account_id,
                sub_work="e-Court Activity",
                purchase_description=ECOURT_PURCHASE_DESCRIPTION,
                description="e-Court Purchase",
                range_from=prior_from,
                range_to_next=date_from,
                ob_date=ob_date,
                seen_ids=seen_purchase_dailies,
            )
            self._append_missing_purpose_purchases(
                txn_rows,
                account_id=account_id,
                sub_work="e-Court Activity",
                purchase_description=ECOURT_PURCHASE_DESCRIPTION,
                description="e-Court Purchase",
                range_from=txn_params["range_from"],
                range_to_next=date_to_next,
                ob_date=ob_date,
                seen_ids=seen_purchase_dailies,
            )
        obc_prior_dr, obc_prior_cr = self._append_orphan_obc_deposits(
            [],
            account_id=account_id,
            range_from=prior_from,
            range_to_next=date_from,
            ob_date=ob_date,
        )
        extra_prior_debit = self._money(extra_prior_debit + obc_prior_dr)
        extra_prior_credit = self._money(extra_prior_credit + obc_prior_cr)
        self._append_orphan_obc_deposits(
            txn_rows,
            account_id=account_id,
            range_from=txn_params["range_from"],
            range_to_next=date_to_next,
            ob_date=ob_date,
        )
        prior_debit = self._money(prior_debit + extra_prior_debit)
        prior_credit = self._money(prior_credit + extra_prior_credit)
        opening = self._apply_bank_running(
            opening, prior_debit, prior_credit, credit_normal=credit_normal
        )
        txn_rows = self._dedupe_bank_txn_rows(txn_rows)

        lines: list[dict[str, Any]] = []
        running = opening
        lines.append(
            {
                "date": date_from.strftime("%d/%m/%Y"),
                "description": "Opening Balance",
                "reference": "OPENING",
                "source": "Bank Master",
                "ledger_kind": "",
                "debit": Decimal("0.00"),
                "credit": Decimal("0.00"),
                "balance": running,
                "kind": "opening",
            }
        )

        def _day_total_row(day: date, debit_sum: Decimal, credit_sum: Decimal) -> dict[str, Any]:
            return {
                "date": day.strftime("%d/%m/%Y"),
                "description": f"{day.strftime('%d/%m/%Y')} Total",
                "reference": "",
                "source": "",
                "ledger_kind": "",
                "debit": debit_sum,
                "credit": credit_sum,
                "balance": None,  # blank on total rows
                "kind": "day_total",
            }

        def _month_total_row(year: int, month: int, debit_sum: Decimal, credit_sum: Decimal) -> dict[str, Any]:
            label_m = f"{month:02d}/{year}"
            return {
                "date": label_m,
                "description": f"{label_m} Total",
                "reference": "",
                "source": "",
                "ledger_kind": "",
                "debit": debit_sum,
                "credit": credit_sum,
                "balance": None,  # blank on total rows
                "kind": "month_total",
            }

        cur_day: date | None = None
        cur_month: tuple[int, int] | None = None
        day_debit = Decimal("0.00")
        day_credit = Decimal("0.00")
        month_debit = Decimal("0.00")
        month_credit = Decimal("0.00")

        for row in txn_rows:
            txn_date = self._to_ledger_date(row["TransactionDate"])
            if txn_date is None:
                continue
            debit = self._money(row["DebitValue"])
            credit = self._money(row["CreditValue"])

            # Day changed → flush previous day total
            if cur_day is not None and txn_date != cur_day:
                lines.append(_day_total_row(cur_day, day_debit, day_credit))
                day_debit = Decimal("0.00")
                day_credit = Decimal("0.00")
                # Month changed after day flush → flush previous month total
                if cur_month is not None and (txn_date.year, txn_date.month) != cur_month:
                    lines.append(
                        _month_total_row(cur_month[0], cur_month[1], month_debit, month_credit)
                    )
                    month_debit = Decimal("0.00")
                    month_credit = Decimal("0.00")

            cur_day = txn_date
            cur_month = (txn_date.year, txn_date.month)
            day_debit = self._money(day_debit + debit)
            day_credit = self._money(day_credit + credit)
            month_debit = self._money(month_debit + debit)
            month_credit = self._money(month_credit + credit)

            running = self._apply_bank_running(
                running, debit, credit, credit_normal=credit_normal
            )
            desc = (row["Description"] or row["Remarks"] or "Bank transaction").strip()
            btid = int(row["JtcsBankTransactionID"] or 0)
            rec_id = int(row["SourceRecordID"]) if row["SourceRecordID"] else None
            if btid > 0:
                ref = f"BT-{btid}"
                if rec_id:
                    ref = f"{ref} / Rec#{rec_id}"
            elif rec_id:
                ref = f"Rec#{rec_id}"
            else:
                ref = ""
            lines.append(
                {
                    "date": txn_date.strftime("%d/%m/%Y"),
                    "description": desc,
                    "reference": ref,
                    "source": (row["SourceTable"] or row["SourceType"] or "").strip(),
                    "source_record_id": rec_id,
                    "bank_transaction_id": btid if btid > 0 else None,
                    "ledger_kind": (row["LedgerKind"] or "").strip(),
                    "debit": debit,
                    "credit": credit,
                    "balance": running,
                    "kind": "txn",
                }
            )

        if cur_day is not None:
            lines.append(_day_total_row(cur_day, day_debit, day_credit))
        if cur_month is not None:
            lines.append(_month_total_row(cur_month[0], cur_month[1], month_debit, month_credit))

        grouped_lines = self._bank_desc_group_lines(
            txn_rows,
            opening=opening,
            date_from=date_from,
            credit_normal=credit_normal,
        )
        pivot = self._bank_pivot_matrix(txn_rows)

        return {
            "kind": "bank",
            "title": "Bank Account Ledger",
            "entity_name": label,
            "entity_id": account_id,
            "safe_name": re.sub(r"[^\w\-]+", "_", bank_name)[:40],
            "meta": [
                ("Account", label),
                ("Account Holder", (account["AccountHolderName"] or "").strip() or "—"),
                ("Chart of Account Group", (account.get("GroupName") or "").strip() or "—"),
                ("Ledger Balance", f"{running:,.2f}"),
                ("Period", f"{date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}"),
            ],
            "headers": [
                "Date",
                "Description",
                "Reference",
                "Source",
                "Ledger Kind",
                "Debit",
                "Credit",
                "Running Balance",
            ],
            "lines": lines,
            "grouped_lines": grouped_lines,
            "pivot": pivot,
            "closing": running,
            "date_from": date_from,
            "date_to": date_to,
        }

    def _bank_pivot_matrix(self, txn_rows: list[Any]) -> dict[str, Any]:
        """Pivot source: Description x Date cells (sheet renders Date rows, Description headers)."""
        cells: dict[tuple[str, date], dict[str, Decimal]] = defaultdict(
            lambda: {"debit": Decimal("0.00"), "credit": Decimal("0.00")}
        )
        dates: set[date] = set()
        descs: set[str] = set()

        for row in txn_rows:
            txn_date = self._to_ledger_date(row["TransactionDate"])
            if txn_date is None:
                continue
            desc = (row["Description"] or row["Remarks"] or "Bank transaction").strip()
            dates.add(txn_date)
            descs.add(desc)
            key = (desc, txn_date)
            cells[key]["debit"] = self._money(cells[key]["debit"] + self._money(row["DebitValue"]))
            cells[key]["credit"] = self._money(
                cells[key]["credit"] + self._money(row["CreditValue"])
            )

        return {
            "dates": sorted(dates),
            "descriptions": sorted(descs, key=str.lower),
            "cells": dict(cells),
        }

    def _bank_desc_group_lines(
        self,
        txn_rows: list[Any],
        *,
        opening: Decimal,
        date_from: date,
        credit_normal: bool = False,
    ) -> list[dict[str, Any]]:
        """Sheet-2 view: sort by date then description; group totals per date+description."""
        lines: list[dict[str, Any]] = []
        running = opening
        lines.append(
            {
                "date": date_from.strftime("%d/%m/%Y"),
                "description": "Opening Balance",
                "reference": "OPENING",
                "source": "Bank Master",
                "ledger_kind": "",
                "debit": Decimal("0.00"),
                "credit": Decimal("0.00"),
                "balance": running,
                "kind": "opening",
            }
        )

        def _row_desc(row: Any) -> str:
            return (row["Description"] or row["Remarks"] or "Bank transaction").strip()

        def _group_total_row(
            day: date, desc: str, debit_sum: Decimal, credit_sum: Decimal
        ) -> dict[str, Any]:
            day_label = day.strftime("%d/%m/%Y")
            return {
                "date": day_label,
                "description": f"{day_label} - {desc} Total",
                "reference": "",
                "source": "",
                "ledger_kind": "",
                "debit": debit_sum,
                "credit": credit_sum,
                "balance": None,
                "kind": "desc_total",
            }

        sorted_rows = sorted(
            [r for r in txn_rows if r["TransactionDate"] is not None],
            key=lambda r: (
                self._to_ledger_date(r["TransactionDate"]) or date.min,
                _row_desc(r).lower(),
                int(r["JtcsBankTransactionID"] or 0),
            ),
        )

        cur_key: tuple[date, str] | None = None
        group_debit = Decimal("0.00")
        group_credit = Decimal("0.00")

        for row in sorted_rows:
            txn_date = self._to_ledger_date(row["TransactionDate"])
            if txn_date is None:
                continue
            desc = _row_desc(row)
            key = (txn_date, desc)
            debit = self._money(row["DebitValue"])
            credit = self._money(row["CreditValue"])

            if cur_key is not None and key != cur_key:
                lines.append(
                    _group_total_row(cur_key[0], cur_key[1], group_debit, group_credit)
                )
                group_debit = Decimal("0.00")
                group_credit = Decimal("0.00")

            cur_key = key
            group_debit = self._money(group_debit + debit)
            group_credit = self._money(group_credit + credit)
            running = self._apply_bank_running(
                running, debit, credit, credit_normal=credit_normal
            )

            ref = f"BT-{row['JtcsBankTransactionID']}"
            if row["SourceRecordID"]:
                ref = f"{ref} / Rec#{row['SourceRecordID']}"
            lines.append(
                {
                    "date": txn_date.strftime("%d/%m/%Y"),
                    "description": desc,
                    "reference": ref,
                    "source": (row["SourceTable"] or row["SourceType"] or "").strip(),
                    "source_record_id": int(row["SourceRecordID"]) if row["SourceRecordID"] else None,
                    "bank_transaction_id": int(row["JtcsBankTransactionID"]),
                    "ledger_kind": (row["LedgerKind"] or "").strip(),
                    "debit": debit,
                    "credit": credit,
                    "balance": running,
                    "kind": "txn",
                }
            )

        if cur_key is not None:
            lines.append(
                _group_total_row(cur_key[0], cur_key[1], group_debit, group_credit)
            )
        return lines

    def _customer_coa_ledger_key(self, customer_id: int) -> str | None:
        """Chart of Account ledger key for a customer (used by Bank/Cash Electronic Transfer)."""
        if not db.session.execute(
            text("SELECT OBJECT_ID(N'dbo.ChartOfAccountMaster', N'U')")
        ).scalar():
            return None
        if not db.session.execute(
            text(
                "SELECT CASE WHEN COL_LENGTH(N'dbo.ChartOfAccountMaster', N'CustomerID') "
                "IS NULL THEN 0 ELSE 1 END"
            )
        ).scalar():
            return None
        account_id = db.session.execute(
            text(
                """
                SELECT TOP 1 AccountID
                FROM dbo.ChartOfAccountMaster
                WHERE CustomerID = :customer_id
                  AND ISNULL(IsActive, 1) = 1
                ORDER BY AccountID ASC
                """
            ),
            {"customer_id": customer_id},
        ).scalar()
        if not account_id:
            return None
        return f"coa-{int(account_id)}"

    def _customer_obc_ledger_parts(
        self,
        customer_id: int,
        *,
        date_from: date,
        date_to: date,
        ob_date: date | None,
    ) -> tuple[Decimal, Decimal, list[dict[str, Any]]]:
        """Include Other Bank/Cash Electronic Transfer legs on the customer CoA."""
        empty = (Decimal("0.00"), Decimal("0.00"), [])
        if not db.session.execute(
            text("SELECT OBJECT_ID(N'dbo.OthersBankCashTransaction', N'U')")
        ).scalar():
            return empty
        has_keys = bool(
            db.session.execute(
                text(
                    "SELECT CASE WHEN COL_LENGTH(N'dbo.OthersBankCashTransaction', "
                    "N'CreditLedgerKey') IS NULL THEN 0 ELSE 1 END"
                )
            ).scalar()
        )
        if not has_keys:
            return empty

        coa_key = self._customer_coa_ledger_key(customer_id)
        if not coa_key:
            return empty

        prior_date_sql = "AND e.WorkDate < :date_from"
        prior_params: dict[str, Any] = {
            "coa_key": coa_key,
            "date_from": date_from,
        }
        if ob_date is not None:
            prior_date_sql += " AND e.WorkDate > :ob_date"
            prior_params["ob_date"] = ob_date

        prior = db.session.execute(
            text(
                f"""
                SELECT
                    ISNULL(SUM(CASE
                        WHEN e.DebitLedgerKey = :coa_key THEN e.Amount ELSE 0 END), 0) AS billed,
                    ISNULL(SUM(CASE
                        WHEN e.CreditLedgerKey = :coa_key THEN e.Amount ELSE 0 END), 0) AS received
                FROM dbo.OthersBankCashTransaction e
                WHERE ISNULL(e.IsActive, 1) = 1
                  AND (
                        e.CreditLedgerKey = :coa_key
                     OR e.DebitLedgerKey = :coa_key
                  )
                  {prior_date_sql}
                """
            ),
            prior_params,
        ).mappings().first()
        prior_billed = self._money(prior["billed"] if prior else 0)
        prior_received = self._money(prior["received"] if prior else 0)

        raw_rows = db.session.execute(
            text(
                """
                SELECT
                    e.EntryID,
                    e.WorkDate,
                    e.VoucherNo,
                    e.Purpose,
                    e.Remarks,
                    e.Amount,
                    e.CreditLedgerKey,
                    e.DebitLedgerKey
                FROM dbo.OthersBankCashTransaction e
                WHERE ISNULL(e.IsActive, 1) = 1
                  AND (
                        e.CreditLedgerKey = :coa_key
                     OR e.DebitLedgerKey = :coa_key
                  )
                  AND e.WorkDate >= :date_from
                  AND e.WorkDate <= :date_to
                ORDER BY e.WorkDate ASC, e.EntryID ASC
                """
            ),
            {
                "coa_key": coa_key,
                "date_from": date_from,
                "date_to": date_to,
            },
        ).mappings().all()

        rows: list[dict[str, Any]] = []
        for r in raw_rows:
            amount = self._money(r["Amount"])
            debit_key = (r.get("DebitLedgerKey") or "").strip()
            credit_key = (r.get("CreditLedgerKey") or "").strip()
            # Debit leg (Money In) → customer Debit column; Credit leg (Money Out) → Credit.
            is_debit_side = debit_key == coa_key
            is_credit_side = credit_key == coa_key
            if not is_debit_side and not is_credit_side:
                continue
            purpose = (r.get("Purpose") or "").strip() or "Electronic Transfer"
            voucher = (r.get("VoucherNo") or "").strip()
            direction = "Money In" if is_debit_side else "Money Out"
            counter_key = credit_key if is_debit_side else debit_key
            counter_label = self._obc_ledger_key_label(counter_key)
            desc = f"Other Bank/Cash — {purpose} ({direction})"
            if counter_label:
                desc = f"{desc} · {counter_label}"
            rows.append(
                {
                    "TransactionID": -int(r["EntryID"] or 0),
                    "TransactionDate": r.get("WorkDate"),
                    "WorkType": "Others",
                    "SubWorkType": f"Other Bank/Cash Transactions - {purpose}",
                    "StampID": None,
                    "ReferenceNo": voucher,
                    "Description": desc,
                    "Remarks": (r.get("Remarks") or "").strip(),
                    "SaleAmount": amount if is_debit_side else Decimal("0.00"),
                    "IncomeAmount": Decimal("0.00"),
                    "BankDebit": amount if is_credit_side else Decimal("0.00"),
                    "PaymentTotal": Decimal("0.00"),
                    "ReceiptAmount": amount if is_credit_side else Decimal("0.00"),
                    "ObcEntryID": int(r["EntryID"] or 0),
                }
            )

        return prior_billed, prior_received, rows

    def _obc_ledger_key_label(self, ledger_key: str) -> str:
        key = (ledger_key or "").strip()
        if not key:
            return ""
        if key.startswith("bank-"):
            try:
                bank_id = int(key.split("-", 1)[1])
            except (TypeError, ValueError):
                return key
            row = db.session.execute(
                text(
                    """
                    SELECT BankName, ISNULL(MaskedAccountNumber, AccountNumber) AS AccRef
                    FROM dbo.JtcsBankAccountMaster
                    WHERE JtcsBankAccountID = :bid
                    """
                ),
                {"bid": bank_id},
            ).mappings().first()
            if not row:
                return key
            name = (row.get("BankName") or "").strip()
            ref = (row.get("AccRef") or "").strip()
            return f"{name} {ref}".strip() or key
        if key.startswith("coa-"):
            try:
                aid = int(key.split("-", 1)[1])
            except (TypeError, ValueError):
                return key
            row = db.session.execute(
                text(
                    """
                    SELECT AccountName
                    FROM dbo.ChartOfAccountMaster
                    WHERE AccountID = :aid
                    """
                ),
                {"aid": aid},
            ).mappings().first()
            if not row:
                return key
            return (row.get("AccountName") or "").strip() or key
        return key

    def _customer_ledger_data(
        self,
        customer_id: int,
        *,
        date_from: date | None,
        date_to: date | None,
    ) -> dict[str, Any]:
        has_ob = bool(
            db.session.execute(
                text(
                    "SELECT CASE WHEN COL_LENGTH(N'dbo.CustomerMaster', N'OpeningBalance') "
                    "IS NULL THEN 0 ELSE 1 END"
                )
            ).scalar()
        )
        customer_sql = """
            SELECT CustomerID, CustomerName, MobileNumber, PANNumber
            FROM dbo.CustomerMaster
            WHERE CustomerID = :customer_id
        """
        if has_ob:
            customer_sql = """
                SELECT CustomerID, CustomerName, MobileNumber, PANNumber,
                       ISNULL(OpeningBalance, 0) AS OpeningBalance,
                       OpeningBalanceDate,
                       OpeningBalanceDrCr,
                       CustomerGroup
                FROM dbo.CustomerMaster
                WHERE CustomerID = :customer_id
            """
        else:
            customer_sql = """
                SELECT CustomerID, CustomerName, MobileNumber, PANNumber, CustomerGroup
                FROM dbo.CustomerMaster
                WHERE CustomerID = :customer_id
            """
        customer = db.session.execute(
            text(customer_sql),
            {"customer_id": customer_id},
        ).mappings().first()
        if customer is None:
            raise ValueError("Customer not found.")

        date_from, date_to = self._resolve_period(date_from, date_to)
        opening = Decimal("0.00")
        ob_date = None
        if has_ob:
            ob_date = customer["OpeningBalanceDate"]
            ob_amount = self._money(customer["OpeningBalance"])
            ob_type = (customer["OpeningBalanceDrCr"] or "Dr").strip()
            # Dr = receivable (+), Cr = advance / credit (-)
            signed_ob = ob_amount if ob_type.upper().startswith("D") else -ob_amount
            if ob_amount != 0 and (ob_date is None or ob_date <= date_from):
                opening = signed_ob

        prior_params = {"customer_id": customer_id, "date_from": date_from}
        prior_date_sql = "AND d.TransactionDate < :date_from"
        if ob_date is not None:
            prior_date_sql += " AND d.TransactionDate > :ob_date"
            prior_params["ob_date"] = ob_date

        prior = db.session.execute(
            text(
                f"""
                SELECT
                    ISNULL(SUM(x.billed), 0) AS billed,
                    ISNULL(SUM(x.received), 0) AS received
                FROM (
                    SELECT
                        ISNULL(d.SaleAmount, 0) + ISNULL(d.IncomeAmount, 0) AS billed,
                        {sql_customer_receipt_expr("d", "b")} AS received
                    FROM dbo.JTCSDailyTransaction d
                    LEFT JOIN dbo.JtcsBankTransaction b
                        ON b.JtcsBankTransactionID = d.BankTransactionID
                    WHERE d.CustomerID = :customer_id
                      AND d.Status = N'Posted'
                      {prior_date_sql}
                ) x
                """
            ),
            prior_params,
        ).mappings().first()
        prior_billed = self._money(prior["billed"] if prior else 0)
        prior_received = self._money(prior["received"] if prior else 0)

        # Unpaid Followup Tally bills (ITR/GST/etc.) live on FollowupEntryMaster
        # until Payment Received creates JTCSDailyTransaction — include them so
        # Ledger Report matches Followup billing.
        prior_followup_billed = Decimal("0.00")
        followup_rows: list[Any] = []
        try:
            fu_prior_sql = """
                AND ISNULL(f.BillDate, f.WorkDate) < :date_from
            """
            fu_prior_params = {
                "customer_id": customer_id,
                "date_from": date_from,
            }
            if ob_date is not None:
                fu_prior_sql += " AND ISNULL(f.BillDate, f.WorkDate) > :ob_date"
                fu_prior_params["ob_date"] = ob_date
            prior_followup_billed = self._money(
                db.session.execute(
                    text(
                        f"""
                        SELECT ISNULL(SUM(ISNULL(f.BillAmount, 0)), 0)
                        FROM dbo.FollowupEntryMaster f
                        WHERE f.CustomerID = :customer_id
                          AND ISNULL(f.IsActive, 1) = 1
                          AND f.BillNo IS NOT NULL
                          AND LTRIM(RTRIM(f.BillNo)) <> N''
                          AND ISNULL(f.BillAmount, 0) > 0
                          {fu_prior_sql}
                          {sql_unpaid_followup_exclusion()}
                        """
                    ),
                    fu_prior_params,
                ).scalar()
            )
            followup_rows = list(
                db.session.execute(
                    text(
                        f"""
                        SELECT
                            CAST(NULL AS INT) AS TransactionID,
                            ISNULL(f.BillDate, f.WorkDate) AS TransactionDate,
                            f.ModuleCode AS WorkType,
                            CONCAT(f.ModuleCode, N' Followup') AS SubWorkType,
                            CAST(NULL AS INT) AS StampID,
                            f.BillNo AS ReferenceNo,
                            CONCAT(
                                f.ModuleCode, N' Followup — ',
                                LTRIM(RTRIM(f.BillNo))
                            ) AS Description,
                            CAST(NULL AS NVARCHAR(500)) AS Remarks,
                            ISNULL(f.BillAmount, 0) AS SaleAmount,
                            CAST(0 AS DECIMAL(18, 2)) AS IncomeAmount,
                            CAST(0 AS DECIMAL(18, 2)) AS BankDebit,
                            CAST(0 AS DECIMAL(18, 2)) AS PaymentTotal,
                            CAST(0 AS DECIMAL(18, 2)) AS ReceiptAmount
                        FROM dbo.FollowupEntryMaster f
                        WHERE f.CustomerID = :customer_id
                          AND ISNULL(f.IsActive, 1) = 1
                          AND f.BillNo IS NOT NULL
                          AND LTRIM(RTRIM(f.BillNo)) <> N''
                          AND ISNULL(f.BillAmount, 0) > 0
                          AND ISNULL(f.BillDate, f.WorkDate) >= :date_from
                          AND ISNULL(f.BillDate, f.WorkDate) <= :date_to
                          {sql_unpaid_followup_exclusion()}
                        """
                    ),
                    {
                        "customer_id": customer_id,
                        "date_from": date_from,
                        "date_to": date_to,
                    },
                ).mappings().all()
            )
        except Exception:
            db.session.rollback()
            prior_followup_billed = Decimal("0.00")
            followup_rows = []

        # Other Bank/Cash / Electronic Transfer legs posted to this customer's CoA.
        prior_obc_billed = Decimal("0.00")
        prior_obc_received = Decimal("0.00")
        obc_rows: list[Any] = []
        try:
            prior_obc_billed, prior_obc_received, obc_rows = self._customer_obc_ledger_parts(
                customer_id,
                date_from=date_from,
                date_to=date_to,
                ob_date=ob_date,
            )
        except Exception:
            db.session.rollback()
            prior_obc_billed = Decimal("0.00")
            prior_obc_received = Decimal("0.00")
            obc_rows = []

        opening = self._money(
            opening
            + prior_billed
            + prior_followup_billed
            + prior_obc_billed
            - prior_received
            - prior_obc_received
        )

        rows = list(
            db.session.execute(
                text(
                    f"""
                    SELECT
                        d.TransactionID, d.TransactionDate, d.WorkType, d.SubWorkType,
                        d.StampID, d.ReferenceNo, d.Description, d.Remarks,
                        ISNULL(d.SaleAmount, 0) AS SaleAmount,
                        ISNULL(d.IncomeAmount, 0) AS IncomeAmount,
                        ISNULL(b.Debit, 0) AS BankDebit,
                        (
                            SELECT ISNULL(SUM(p.Amount), 0)
                            FROM dbo.JTCSDailyTransactionPayment p
                            WHERE p.TransactionID = d.TransactionID
                        ) AS PaymentTotal,
                        {sql_customer_receipt_expr("d", "b")} AS ReceiptAmount
                    FROM dbo.JTCSDailyTransaction d
                    LEFT JOIN dbo.JtcsBankTransaction b
                        ON b.JtcsBankTransactionID = d.BankTransactionID
                    WHERE d.CustomerID = :customer_id
                      AND d.Status = N'Posted'
                      AND d.TransactionDate >= :date_from
                      AND d.TransactionDate <= :date_to
                    ORDER BY d.TransactionDate ASC, d.TransactionID ASC
                    """
                ),
                {"customer_id": customer_id, "date_from": date_from, "date_to": date_to},
            ).mappings().all()
        )
        if followup_rows:
            rows.extend(followup_rows)
        if obc_rows:
            rows.extend(obc_rows)
        if followup_rows or obc_rows:
            rows.sort(
                key=lambda r: (
                    r.get("TransactionDate") or date.min,
                    int(r.get("TransactionID") or 0),
                )
            )

        name = (customer["CustomerName"] or f"Customer {customer_id}").strip()
        chart_group_name = ""
        try:
            chart_row = db.session.execute(
                text(
                    """
                    SELECT TOP 1 g.GroupName
                    FROM dbo.ChartOfAccountMaster a
                    INNER JOIN dbo.ChartOfGroupMaster g ON g.GroupID = a.GroupID
                    WHERE a.CustomerID = :cid AND ISNULL(a.IsActive, 1) = 1
                    ORDER BY a.AccountID
                    """
                ),
                {"cid": customer_id},
            ).mappings().first()
            if chart_row:
                chart_group_name = (chart_row.get("GroupName") or "").strip()
        except Exception:
            db.session.rollback()
        customer_group = ""
        try:
            customer_group = (customer.get("CustomerGroup") or "").strip()
        except Exception:
            customer_group = ""
        lines: list[dict[str, Any]] = []
        running = opening
        lines.append(
            {
                "date": date_from.strftime("%d/%m/%Y"),
                "bill": "",
                "work": "",
                "description": "Opening Balance",
                "debit": Decimal("0.00"),
                "credit": Decimal("0.00"),
                "balance": running,
                "kind": "opening",
            }
        )
        for row in rows:
            billed = self._money(row["SaleAmount"]) + self._money(row.get("IncomeAmount"))
            if "ReceiptAmount" in row and row["ReceiptAmount"] is not None:
                receipt = self._money(row["ReceiptAmount"])
            else:
                receipt = self._money(row.get("PaymentTotal"))
                if receipt == 0:
                    receipt = self._money(row.get("BankDebit"))
            work = (row["WorkType"] or "").strip()
            sub = (row["SubWorkType"] or "").strip()
            if sub:
                work = f"{work} / {sub}" if work else sub
            txn_date = row["TransactionDate"]
            ref = (row["ReferenceNo"] or "").strip()
            if not ref and row.get("TransactionID"):
                ref = f"TXN-{row['TransactionID']}"
            raw_desc = (row["Description"] or row.get("Remarks") or "").strip()
            sub_l = sub.lower()
            desc_l = raw_desc.lower()
            is_receipt_row = "followup receipt" in sub_l or desc_l.startswith("payment received") or desc_l.startswith("advance payment")
            is_advance = desc_l.startswith("advance payment") or (is_receipt_row and billed == 0)
            date_str = txn_date.strftime("%d/%m/%Y") if txn_date else ""
            base = {
                "date": date_str,
                "bill": ref,
                "work": work,
                "kind": "txn",
                "transaction_id": row.get("TransactionID"),
                "work_type": row.get("WorkType"),
                "sub_work_type": row.get("SubWorkType"),
                "stamp_id": row.get("StampID"),
                "obc_entry_id": row.get("ObcEntryID"),
            }
            # Never put Debit and Credit of the same payment on one customer-ledger line.
            if billed > 0 and receipt > 0:
                running = self._money(running + billed)
                invoice_desc = raw_desc
                if is_receipt_row or "payment received" in desc_l:
                    invoice_desc = f"Invoice / Sale / Service — {ref}".strip(" —")
                lines.append({**base, "description": invoice_desc or "Invoice / Sale / Service", "debit": billed, "credit": Decimal("0.00"), "balance": running})
                running = self._money(running - receipt)
                pay_desc = "Advance Payment" if is_advance else "Payment Received"
                if ref:
                    pay_desc = f"{pay_desc} — {ref}"
                lines.append({**base, "description": pay_desc, "debit": Decimal("0.00"), "credit": receipt, "balance": running})
                continue
            if billed > 0:
                running = self._money(running + billed)
                desc = raw_desc or "Invoice / Sale / Service"
                lines.append({**base, "description": desc, "debit": billed, "credit": Decimal("0.00"), "balance": running})
                continue
            if receipt > 0:
                running = self._money(running - receipt)
                if is_advance:
                    desc = raw_desc or (f"Advance Payment — {ref}" if ref else "Advance Payment")
                else:
                    desc = raw_desc or (f"Payment Received — {ref}" if ref else "Payment Received")
                lines.append({**base, "description": desc, "debit": Decimal("0.00"), "credit": receipt, "balance": running})

        return {
            "kind": "customer",
            "title": "Customer Ledger",
            "entity_name": name,
            "entity_id": customer_id,
            "safe_name": re.sub(r"[^\w\-]+", "_", name)[:40],
            "meta": [
                ("Customer", name),
                ("Customer ID", str(customer_id)),
                ("Chart of Account Group", chart_group_name or "—"),
                ("Customer Group", customer_group or "—"),
                ("Ledger Balance", f"{running:,.2f}"),
                ("Mobile", (customer["MobileNumber"] or "").strip() or "—"),
                ("PAN", (customer["PANNumber"] or "").strip() or "—"),
                ("Period", f"{date_from.strftime('%d/%m/%Y')} to {date_to.strftime('%d/%m/%Y')}"),
            ],
            "headers": [
                "Date",
                "Bill / Ref No.",
                "Work Type",
                "Description",
                "Debit (Bill)",
                "Credit (Receipt)",
                "Running Balance",
            ],
            "lines": lines,
            "closing": running,
            "date_from": date_from,
            "date_to": date_to,
        }

    def _write_xlsx_ledger_sheet(
        self, ws, data: dict[str, Any], lines: list[dict[str, Any]]
    ) -> None:
        thin = Border(
            left=Side(style="thin", color=COLOR_BORDER),
            right=Side(style="thin", color=COLOR_BORDER),
            top=Side(style="thin", color=COLOR_BORDER),
            bottom=Side(style="thin", color=COLOR_BORDER),
        )
        title_font = Font(name="Calibri", bold=True, size=18, color=COLOR_HEADER_FG)
        brand_font = Font(name="Calibri", bold=True, size=11, color="D5F5E3")
        meta_label_font = Font(name="Calibri", bold=True, size=10, color=COLOR_META_LABEL)
        meta_value_font = Font(name="Calibri", size=10, color="1C2833")
        header_font = Font(name="Calibri", bold=True, size=10, color=COLOR_HEADER_FG)
        cell_font = Font(name="Calibri", size=10, color="1C2833")
        opening_font = Font(name="Calibri", bold=True, size=10, color="7D6608")
        closing_font = Font(name="Calibri", bold=True, size=11, color="0E6655")
        debit_font = Font(name="Calibri", size=10, color=COLOR_DEBIT)
        credit_font = Font(name="Calibri", size=10, color=COLOR_CREDIT)
        balance_font = Font(name="Calibri", bold=True, size=10, color=COLOR_BALANCE)

        fill_title = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        fill_brand = PatternFill("solid", fgColor=COLOR_TEAL)
        fill_meta = PatternFill("solid", fgColor=COLOR_META_BG)
        fill_header = PatternFill("solid", fgColor=COLOR_NAVY)
        fill_alt = PatternFill("solid", fgColor=COLOR_ALT_ROW)
        fill_opening = PatternFill("solid", fgColor=COLOR_OPENING)
        fill_closing = PatternFill("solid", fgColor=COLOR_CLOSING)
        fill_day_total = PatternFill("solid", fgColor="FDEBD0")
        fill_month_total = PatternFill("solid", fgColor="D6EAF8")
        fill_desc_total = PatternFill("solid", fgColor="F5CBA7")
        fill_white = PatternFill("solid", fgColor="FFFFFF")
        total_font = Font(name="Calibri", bold=True, size=10, color="6E2C00")
        month_total_font = Font(name="Calibri", bold=True, size=10, color="1A5276")
        desc_total_font = Font(name="Calibri", bold=True, size=10, color="6E2C00")

        col_count = len(data["headers"])

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"] = "Joshi Tax Consultancy & Services"
        ws["A1"].font = brand_font
        ws["A1"].fill = fill_brand
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws["A2"] = data["title"]
        ws["A2"].font = title_font
        ws["A2"].fill = fill_title
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 28

        row_idx = 4
        for label, value in data["meta"]:
            ws.cell(row=row_idx, column=1, value=label).font = meta_label_font
            ws.cell(row=row_idx, column=1).fill = fill_meta
            ws.cell(row=row_idx, column=1).border = thin
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=min(4, col_count))
            cell = ws.cell(row=row_idx, column=2, value=value)
            cell.font = meta_value_font
            cell.fill = fill_meta
            cell.border = thin
            for c in range(3, min(4, col_count) + 1):
                ws.cell(row=row_idx, column=c).fill = fill_meta
                ws.cell(row=row_idx, column=c).border = thin
            row_idx += 1

        row_idx += 1
        header_row = row_idx
        for col, header in enumerate(data["headers"], start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        ws.row_dimensions[header_row].height = 24

        money_cols_bank = {6, 7, 8}
        money_cols_cust = {5, 6, 7}
        money_cols = money_cols_bank if data["kind"] == "bank" else money_cols_cust

        data_start = header_row + 1
        for i, line in enumerate(lines):
            r = data_start + i
            line_kind = line.get("kind") or "txn"
            bal = line.get("balance")
            if data["kind"] == "bank":
                values = [
                    line["date"],
                    line["description"],
                    line["reference"],
                    line["source"],
                    line["ledger_kind"],
                    float(line["debit"]),
                    float(line["credit"]),
                    "" if bal is None else float(bal),
                ]
            else:
                values = [
                    line["date"],
                    line["bill"],
                    line["work"],
                    line["description"],
                    float(line["debit"]),
                    float(line["credit"]),
                    "" if bal is None else float(bal),
                ]

            if line_kind == "opening":
                row_fill = fill_opening
                row_font = opening_font
            elif line_kind == "day_total":
                row_fill = fill_day_total
                row_font = total_font
            elif line_kind == "month_total":
                row_fill = fill_month_total
                row_font = month_total_font
            elif line_kind == "desc_total":
                row_fill = fill_desc_total
                row_font = desc_total_font
            else:
                row_fill = fill_alt if i % 2 else fill_white
                row_font = None

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.border = thin
                cell.fill = row_fill
                if row_font is not None:
                    cell.font = row_font
                    if col in money_cols and value != "":
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col in money_cols:
                    if data["kind"] == "bank":
                        if col == 6:
                            cell.font = debit_font
                        elif col == 7:
                            cell.font = credit_font
                        else:
                            cell.font = balance_font
                    else:
                        if col == 5:
                            cell.font = debit_font
                        elif col == 6:
                            cell.font = credit_font
                        else:
                            cell.font = balance_font
                    if value != "":
                        cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        close_row = data_start + len(lines) + 1
        ws.cell(row=close_row, column=1, value="Closing Balance").font = closing_font
        ws.cell(row=close_row, column=1).fill = fill_closing
        ws.cell(row=close_row, column=1).border = thin
        for col in range(2, col_count):
            cell = ws.cell(row=close_row, column=col, value="")
            cell.fill = fill_closing
            cell.border = thin
        bal_cell = ws.cell(row=close_row, column=col_count, value=float(data["closing"]))
        bal_cell.font = closing_font
        bal_cell.fill = fill_closing
        bal_cell.border = thin
        bal_cell.number_format = "#,##0.00"
        bal_cell.alignment = Alignment(horizontal="right")

        footer_row = close_row + 2
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=col_count)
        ws.cell(
            row=footer_row,
            column=1,
            value=f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} · JTCS ERP · Confidential",
        ).font = Font(name="Calibri", italic=True, size=9, color="7F8C8D")

        if data["kind"] == "bank":
            widths = [12, 36, 18, 16, 12, 12, 12, 14]
        else:
            widths = [12, 14, 22, 34, 12, 14, 14]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = f"A{header_row + 1}"
        ws.print_title_rows = f"1:{header_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    def _styled_xlsx(self, data: dict[str, Any]) -> tuple[bytes, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger" if data["kind"] == "bank" else "Customer Ledger"
        self._write_xlsx_ledger_sheet(ws, data, data["lines"])

        # Bank only: extra sheet — date + description sort with group totals
        if data["kind"] == "bank" and data.get("grouped_lines"):
            ws2 = wb.create_sheet("Date-Description")
            grouped_data = {
                **data,
                "title": "Bank Account Ledger (Date + Description)",
            }
            self._write_xlsx_ledger_sheet(ws2, grouped_data, data["grouped_lines"])

        # Bank only: pivot sheet — Description rows × Date columns (Debit/Credit)
        if data["kind"] == "bank" and data.get("pivot"):
            ws3 = wb.create_sheet("Pivot")
            self._write_xlsx_pivot_sheet(ws3, data, data["pivot"])

        buffer = io.BytesIO()
        wb.save(buffer)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = "Bank_Ledger" if data["kind"] == "bank" else "Customer_Ledger"
        filename = f"{prefix}_{data['safe_name']}_{data['entity_id']}_{stamp}.xlsx"
        return buffer.getvalue(), filename

    def _write_xlsx_pivot_sheet(
        self, ws, data: dict[str, Any], pivot: dict[str, Any]
    ) -> None:
        """Pivot matrix: row=Date, column=Description (header), values=Debit & Credit."""
        dates: list[date] = pivot.get("dates") or []
        descriptions: list[str] = pivot.get("descriptions") or []
        cells: dict[tuple[str, date], dict[str, Decimal]] = pivot.get("cells") or {}

        thin = Border(
            left=Side(style="thin", color=COLOR_BORDER),
            right=Side(style="thin", color=COLOR_BORDER),
            top=Side(style="thin", color=COLOR_BORDER),
            bottom=Side(style="thin", color=COLOR_BORDER),
        )
        brand_font = Font(name="Calibri", bold=True, size=11, color="D5F5E3")
        title_font = Font(name="Calibri", bold=True, size=16, color=COLOR_HEADER_FG)
        meta_label_font = Font(name="Calibri", bold=True, size=10, color=COLOR_META_LABEL)
        meta_value_font = Font(name="Calibri", size=10, color="1C2833")
        header_font = Font(name="Calibri", bold=True, size=9, color=COLOR_HEADER_FG)
        cell_font = Font(name="Calibri", size=9, color="1C2833")
        total_font = Font(name="Calibri", bold=True, size=9, color="1A5276")
        debit_font = Font(name="Calibri", size=9, color=COLOR_DEBIT)
        credit_font = Font(name="Calibri", size=9, color=COLOR_CREDIT)

        fill_brand = PatternFill("solid", fgColor=COLOR_TEAL)
        fill_title = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        fill_meta = PatternFill("solid", fgColor=COLOR_META_BG)
        fill_header = PatternFill("solid", fgColor=COLOR_NAVY)
        fill_sub = PatternFill("solid", fgColor="2874A6")
        fill_alt = PatternFill("solid", fgColor=COLOR_ALT_ROW)
        fill_white = PatternFill("solid", fgColor="FFFFFF")
        fill_total = PatternFill("solid", fgColor="D6EAF8")
        fill_grand = PatternFill("solid", fgColor="D5F5E3")

        # Each description = 2 columns (Debit, Credit); + Date; + Grand Total Debit/Credit
        col_count = 1 + (len(descriptions) * 2) + 2
        if col_count < 4:
            col_count = 4

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"] = "Joshi Tax Consultancy & Services"
        ws["A1"].font = brand_font
        ws["A1"].fill = fill_brand
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws["A2"] = "Bank Account Ledger - Pivot (Date x Description)"
        ws["A2"].font = title_font
        ws["A2"].fill = fill_title
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 26

        row_idx = 4
        for label, value in data["meta"]:
            ws.cell(row=row_idx, column=1, value=label).font = meta_label_font
            ws.cell(row=row_idx, column=1).fill = fill_meta
            ws.cell(row=row_idx, column=1).border = thin
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=min(4, col_count))
            cell = ws.cell(row=row_idx, column=2, value=value)
            cell.font = meta_value_font
            cell.fill = fill_meta
            cell.border = thin
            for c in range(3, min(4, col_count) + 1):
                ws.cell(row=row_idx, column=c).fill = fill_meta
                ws.cell(row=row_idx, column=c).border = thin
            row_idx += 1

        row_idx += 1
        desc_header_row = row_idx
        sub_header_row = row_idx + 1

        # Corner: Date (row labels)
        corner = ws.cell(row=desc_header_row, column=1, value="Date")
        corner.font = header_font
        corner.fill = fill_header
        corner.alignment = Alignment(horizontal="center", vertical="center")
        corner.border = thin
        ws.merge_cells(
            start_row=desc_header_row,
            start_column=1,
            end_row=sub_header_row,
            end_column=1,
        )
        ws.cell(row=sub_header_row, column=1).fill = fill_header
        ws.cell(row=sub_header_row, column=1).border = thin

        # Description columns (Debit | Credit under each description)
        for i, desc in enumerate(descriptions):
            c0 = 2 + (i * 2)
            c1 = c0 + 1
            ws.merge_cells(
                start_row=desc_header_row,
                start_column=c0,
                end_row=desc_header_row,
                end_column=c1,
            )
            dcell = ws.cell(row=desc_header_row, column=c0, value=desc)
            dcell.font = header_font
            dcell.fill = fill_header
            dcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            dcell.border = thin
            ws.cell(row=desc_header_row, column=c1).fill = fill_header
            ws.cell(row=desc_header_row, column=c1).border = thin

            for offset, label in enumerate(("Debit", "Credit")):
                cell = ws.cell(row=sub_header_row, column=c0 + offset, value=label)
                cell.font = header_font
                cell.fill = fill_sub
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin

        # Grand total columns
        gt0 = 2 + (len(descriptions) * 2)
        gt1 = gt0 + 1
        ws.merge_cells(
            start_row=desc_header_row,
            start_column=gt0,
            end_row=desc_header_row,
            end_column=gt1,
        )
        gcell = ws.cell(row=desc_header_row, column=gt0, value="Grand Total")
        gcell.font = header_font
        gcell.fill = fill_header
        gcell.alignment = Alignment(horizontal="center", vertical="center")
        gcell.border = thin
        ws.cell(row=desc_header_row, column=gt1).fill = fill_header
        ws.cell(row=desc_header_row, column=gt1).border = thin
        for offset, label in enumerate(("Debit", "Credit")):
            cell = ws.cell(row=sub_header_row, column=gt0 + offset, value=label)
            cell.font = header_font
            cell.fill = fill_sub
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

        ws.row_dimensions[desc_header_row].height = 30
        ws.row_dimensions[sub_header_row].height = 18

        # Column totals (per description) for bottom Total row
        col_debit_tot = [Decimal("0.00") for _ in descriptions]
        col_credit_tot = [Decimal("0.00") for _ in descriptions]
        grand_debit = Decimal("0.00")
        grand_credit = Decimal("0.00")

        data_start = sub_header_row + 1
        for r_i, d in enumerate(dates):
            r = data_start + r_i
            row_fill = fill_alt if r_i % 2 else fill_white
            date_cell = ws.cell(row=r, column=1, value=d.strftime("%d/%m/%Y"))
            date_cell.font = cell_font
            date_cell.fill = row_fill
            date_cell.border = thin
            date_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_debit = Decimal("0.00")
            row_credit = Decimal("0.00")
            for i, desc in enumerate(descriptions):
                amounts = cells.get((desc, d)) or {
                    "debit": Decimal("0.00"),
                    "credit": Decimal("0.00"),
                }
                debit = self._money(amounts["debit"])
                credit = self._money(amounts["credit"])
                row_debit = self._money(row_debit + debit)
                row_credit = self._money(row_credit + credit)
                col_debit_tot[i] = self._money(col_debit_tot[i] + debit)
                col_credit_tot[i] = self._money(col_credit_tot[i] + credit)

                c0 = 2 + (i * 2)
                for offset, amount, font in (
                    (0, debit, debit_font),
                    (1, credit, credit_font),
                ):
                    cell = ws.cell(
                        row=r,
                        column=c0 + offset,
                        value=float(amount) if amount != 0 else None,
                    )
                    cell.font = font
                    cell.fill = row_fill
                    cell.border = thin
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

            grand_debit = self._money(grand_debit + row_debit)
            grand_credit = self._money(grand_credit + row_credit)
            for offset, amount, font in (
                (0, row_debit, Font(name="Calibri", bold=True, size=9, color=COLOR_DEBIT)),
                (1, row_credit, Font(name="Calibri", bold=True, size=9, color=COLOR_CREDIT)),
            ):
                cell = ws.cell(
                    row=r,
                    column=gt0 + offset,
                    value=float(amount) if amount != 0 else None,
                )
                cell.font = font
                cell.fill = fill_total
                cell.border = thin
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

        # Bottom Total row
        total_row = data_start + len(dates)
        tcell = ws.cell(row=total_row, column=1, value="Total")
        tcell.font = total_font
        tcell.fill = fill_grand
        tcell.border = thin
        for i, _desc in enumerate(descriptions):
            c0 = 2 + (i * 2)
            for offset, amount in ((0, col_debit_tot[i]), (1, col_credit_tot[i])):
                cell = ws.cell(row=total_row, column=c0 + offset, value=float(amount))
                cell.font = total_font
                cell.fill = fill_grand
                cell.border = thin
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        for offset, amount in ((0, grand_debit), (1, grand_credit)):
            cell = ws.cell(row=total_row, column=gt0 + offset, value=float(amount))
            cell.font = total_font
            cell.fill = fill_grand
            cell.border = thin
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

        footer_row = total_row + 2
        ws.merge_cells(
            start_row=footer_row, start_column=1, end_row=footer_row, end_column=min(6, col_count)
        )
        ws.cell(
            row=footer_row,
            column=1,
            value=f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} · JTCS ERP · Pivot sheet",
        ).font = Font(name="Calibri", italic=True, size=9, color="7F8C8D")

        ws.column_dimensions["A"].width = 12
        for col in range(2, col_count + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.freeze_panes = f"B{data_start}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    def _draw_jtcs_watermark(self, canvas: pdf_canvas.Canvas, doc) -> None:
        """Single light diagonal JTCS watermark (bottom-left → top-right)."""
        canvas.saveState()
        page_w, page_h = doc.pagesize
        angle = math.degrees(math.atan2(page_h, page_w))
        canvas.translate(page_w / 2.0, page_h / 2.0)
        canvas.rotate(angle)
        # One soft background layer only (no double/dark shadow)
        canvas.setFillColor(colors.Color(0.12, 0.55, 0.50, alpha=0.07))
        canvas.setFont("Helvetica-Bold", 180)
        canvas.drawCentredString(0, -45, "JTCS")
        canvas.restoreState()

        # Footer strip
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#148F77"))
        canvas.rect(0, 0, page_w, 14, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18, 4, "Joshi Tax Consultancy & Services · Confidential")
        canvas.drawRightString(page_w - 18, 4, f"Page {doc.page}")
        canvas.restoreState()

    def _styled_pdf(self, data: dict[str, Any]) -> tuple[bytes, str]:
        buffer = io.BytesIO()
        pagesize = A4
        doc = BaseDocTemplate(
            buffer,
            pagesize=pagesize,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=16 * mm,
        )
        frame = Frame(
            doc.leftMargin,
            doc.bottomMargin,
            doc.width,
            doc.height,
            id="normal",
        )
        doc.addPageTemplates(
            [
                PageTemplate(
                    id="ledger",
                    frames=[frame],
                    onPage=self._draw_jtcs_watermark,
                )
            ]
        )

        styles = getSampleStyleSheet()
        brand = ParagraphStyle(
            "Brand",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            textColor=colors.HexColor("#148F77"),
            spaceAfter=2,
        )
        title = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            textColor=colors.HexColor("#1A5276"),
            spaceAfter=8,
        )
        meta_style = ParagraphStyle(
            "Meta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#1C2833"),
            leading=12,
        )
        cell_style = ParagraphStyle(
            "Cell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1C2833"),
        )

        story: list[Any] = [
            Paragraph("Joshi Tax Consultancy &amp; Services", brand),
            Paragraph(data["title"], title),
        ]
        meta_bits = " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(
            f"<b>{label}:</b> {value}" for label, value in data["meta"]
        )
        story.append(Paragraph(meta_bits, meta_style))
        story.append(Spacer(1, 8))

        bold_cell = ParagraphStyle(
            "BoldCell",
            parent=cell_style,
            fontName="Helvetica-Bold",
        )
        story.append(
            self._build_pdf_ledger_table(data, data["lines"], cell_style, bold_cell)
        )

        # Bank only: second section — date + description sort with group totals
        if data["kind"] == "bank" and data.get("grouped_lines"):
            story.append(PageBreak())
            story.append(Paragraph("Joshi Tax Consultancy &amp; Services", brand))
            story.append(
                Paragraph("Bank Account Ledger (Date + Description)", title)
            )
            story.append(Paragraph(meta_bits, meta_style))
            story.append(Spacer(1, 8))
            story.append(
                self._build_pdf_ledger_table(
                    data, data["grouped_lines"], cell_style, bold_cell
                )
            )

        story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} · JTCS ERP",
                ParagraphStyle(
                    "Foot",
                    parent=styles["Normal"],
                    fontSize=8,
                    textColor=colors.HexColor("#7F8C8D"),
                    alignment=TA_CENTER,
                ),
            )
        )

        doc.build(story)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = "Bank_Ledger" if data["kind"] == "bank" else "Customer_Ledger"
        filename = f"{prefix}_{data['safe_name']}_{data['entity_id']}_{stamp}.pdf"
        return buffer.getvalue(), filename

    def _build_pdf_ledger_table(
        self,
        data: dict[str, Any],
        lines: list[dict[str, Any]],
        cell_style: ParagraphStyle,
        bold_cell: ParagraphStyle,
    ) -> Table:
        table_data: list[list[Any]] = [
            [Paragraph(f"<b>{h}</b>", cell_style) for h in data["headers"]]
        ]
        line_kinds: list[str] = ["header"]
        total_kinds = {"day_total", "month_total", "desc_total", "opening"}
        for line in lines:
            line_kind = line.get("kind") or "txn"
            line_kinds.append(line_kind)
            bal = line.get("balance")
            bal_txt = "" if bal is None else f"<b>{self._fmt_money(bal)}</b>"
            style_use = bold_cell if line_kind in total_kinds else cell_style
            if data["kind"] == "bank":
                row = [
                    Paragraph(line["date"], style_use),
                    Paragraph(line["description"] or "—", style_use),
                    Paragraph(line["reference"] or "", style_use),
                    Paragraph(line["source"] or "", style_use),
                    Paragraph(line["ledger_kind"] or "", style_use),
                    Paragraph(self._fmt_money(line["debit"]), style_use),
                    Paragraph(self._fmt_money(line["credit"]), style_use),
                    Paragraph(bal_txt, style_use),
                ]
            else:
                row = [
                    Paragraph(line["date"], style_use),
                    Paragraph(line["bill"] or "", style_use),
                    Paragraph(line["work"] or "", style_use),
                    Paragraph(line["description"] or "—", style_use),
                    Paragraph(self._fmt_money(line["debit"]), style_use),
                    Paragraph(self._fmt_money(line["credit"]), style_use),
                    Paragraph(bal_txt, style_use),
                ]
            table_data.append(row)

        line_kinds.append("closing")
        close_pad = len(data["headers"]) - 2
        close_row = [Paragraph("<b>Closing Balance</b>", cell_style)] + [
            "" for _ in range(close_pad)
        ] + [Paragraph(f"<b>{self._fmt_money(data['closing'])}</b>", cell_style)]
        table_data.append(close_row)

        if data["kind"] == "bank":
            col_widths = [48, 110, 58, 52, 42, 50, 50, 58]
        else:
            col_widths = [48, 55, 78, 118, 52, 52, 58]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D5F5E3")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#5D6D7E")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (-3, 1), (-1, -1), "RIGHT"),
        ]
        for i, kind in enumerate(line_kinds):
            if kind == "header":
                continue
            if kind == "opening":
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FCF3CF")))
            elif kind == "day_total":
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FDEBD0")))
            elif kind == "month_total":
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#D6EAF8")))
            elif kind == "desc_total":
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F5CBA7")))
            elif kind == "closing":
                continue
            elif i % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E8F6F3")))
        table.setStyle(TableStyle(style_cmds))
        return table

    def bank_ledger_preview_data(
        self,
        account_id: int,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> dict[str, Any]:
        """Ledger rows for on-screen HTML preview (Bank Accounts tab)."""
        return self._bank_ledger_data(account_id, date_from=date_from, date_to=date_to)

    def customer_ledger_preview_data(
        self,
        customer_id: int,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> dict[str, Any]:
        """Ledger rows for on-screen HTML preview (Customers tab)."""
        return self._customer_ledger_data(
            customer_id, date_from=date_from, date_to=date_to
        )

    def build_bank_ledger(
        self,
        account_id: int,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
        fmt: str = "xlsx",
    ) -> tuple[bytes, str, str]:
        data = self._bank_ledger_data(account_id, date_from=date_from, date_to=date_to)
        fmt = (fmt or "xlsx").lower().strip()
        if fmt == "pdf":
            content, filename = self._styled_pdf(data)
            return content, filename, "application/pdf"
        content, filename = self._styled_xlsx(data)
        return (
            content,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def build_customer_ledger(
        self,
        customer_id: int,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
        fmt: str = "xlsx",
    ) -> tuple[bytes, str, str]:
        data = self._customer_ledger_data(customer_id, date_from=date_from, date_to=date_to)
        fmt = (fmt or "xlsx").lower().strip()
        if fmt == "pdf":
            content, filename = self._styled_pdf(data)
            return content, filename, "application/pdf"
        content, filename = self._styled_xlsx(data)
        return (
            content,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Back-compat wrappers
    def build_bank_ledger_xlsx(self, account_id: int, **kwargs) -> tuple[bytes, str]:
        data, name, _ = self.build_bank_ledger(account_id, fmt="xlsx", **kwargs)
        return data, name

    def build_customer_ledger_xlsx(self, customer_id: int, **kwargs) -> tuple[bytes, str]:
        data, name, _ = self.build_customer_ledger(customer_id, fmt="xlsx", **kwargs)
        return data, name
