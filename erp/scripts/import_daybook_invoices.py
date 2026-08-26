"""Create Sale invoices from Tally DayBook.xlsx when Tally Bill Number exists in Followup."""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_XLSX = Path(r"C:\Users\USER\Downloads\DayBook.xlsx")
DEFAULT_LOG = Path(r"C:\temp\daybook_invoice_result.csv")
QUARTER_MAP = {
    "Q1": "Q1-Apr-May-Jun",
    "Q2": "Q2-Jul-Aug-Sep",
    "Q3": "Q3-Oct-Nov-Dec",
    "Q4": "Q4-Jan-Feb-Mar",
}


def _norm_bill(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, "f").rstrip("0").rstrip(".")
        return text
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _as_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _as_amount(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _read_sales_rows(xlsx_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Sales Register" not in wb.sheetnames:
        raise ValueError("Excel mein 'Sales Register' sheet nahi mili.")
    ws = wb["Sales Register"]
    rows: list[dict] = []
    header = None
    for raw in ws.iter_rows(values_only=True):
        values = list(raw)
        if header is None:
            header = [str(v or "").strip() for v in values]
            continue
        rec = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
        bill_no = _norm_bill(rec.get("Tally Bill Number"))
        if not bill_no:
            continue
        rows.append(
            {
                "bill_no": bill_no,
                "excel_name": str(rec.get("Customer Name") or "").strip(),
                "excel_date": _as_date(rec.get("Date")),
                "excel_amount": _as_amount(rec.get("Amount")),
            }
        )
    wb.close()
    unique: dict[str, dict] = {}
    for row in rows:
        unique.setdefault(row["bill_no"], row)
    return list(unique.values())


def _map_quarter(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    if value in QUARTER_MAP.values():
        return value
    return QUARTER_MAP.get(value.upper()[:2], "")


def _find_item(items: list[dict], module_code: str) -> dict | None:
    needle = (module_code or "").upper()
    if not needle:
        return None
    for item in items:
        hay = " ".join(
            [
                str(item.get("item_code") or ""),
                str(item.get("item_name") or ""),
                str(item.get("label") or ""),
            ]
        ).upper()
        if needle in hay:
            return item
    return None


def _q2(value) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _split_inclusive(gross, gst_rate, *, intra: bool) -> tuple[Decimal, Decimal]:
    """Excel amount includes GST. Return taxable rate and signed round-off."""
    gross = _q2(gross)
    gst_rate = _q2(gst_rate)
    if gross <= 0:
        return Decimal("0.00"), Decimal("0.00")
    if gst_rate <= 0:
        return gross, Decimal("0.00")
    taxable = _q2(gross * Decimal("100") / (Decimal("100") + gst_rate))
    if intra:
        half = _q2(gst_rate / Decimal("2"))
        gst_total = _q2(taxable * half / Decimal("100")) + _q2(taxable * half / Decimal("100"))
    else:
        gst_total = _q2(taxable * gst_rate / Decimal("100"))
    computed = _q2(taxable + gst_total)
    return taxable, _q2(gross - computed)


def _is_intra(customer_id) -> bool:
    from app.services.gst_invoice_service import GstInvoiceService

    if not customer_id:
        return True
    svc = GstInvoiceService()
    cust = svc._load_customer(int(customer_id))
    company = svc.company_profile()
    place = (cust.get("place_of_supply_code") or "").strip()
    seller = (company.get("state_code") or "05").strip()
    return bool(place) and place == seller


def _round_sign(round_off: Decimal) -> str:
    if round_off > 0:
        return "add"
    if round_off < 0:
        return "sub"
    return ""


def _pick_bank(bank_id: int | None) -> dict:
    from app.services.bank_master_service import BankMasterService

    banks = BankMasterService().list_payment_accounts()
    if not banks:
        raise ValueError("Koi Payment Bank (UPI ID wala) Bank Master mein nahi mila.")
    if bank_id:
        for bank in banks:
            if int(bank.get("account_id") or 0) == int(bank_id):
                return bank
        raise ValueError(f"Payment bank id {bank_id} nahi mili.")
    return banks[0]


def _existing_similar_invoice(customer_id, invoice_date: date, amount: float | None) -> dict | None:
    """Catch invoices saved before TallyBillNo was stored."""
    from sqlalchemy import text

    from app.extensions import db

    if not customer_id or not invoice_date:
        return None
    amt = float(amount or 0)
    row = db.session.execute(
        text(
            """
            SELECT TOP 1 InvoiceNo, CustomerName, InvoiceValue, TaxableValue
            FROM dbo.GstInvoice
            WHERE ISNULL(VoucherType, N'SALE') = N'SALE'
              AND CustomerID = :cid
              AND InvoiceDate = :dt
              AND (
                (ISNULL(TallyBillNo, N'') <> N'')
                OR ABS(CAST(ISNULL(TaxableValue, 0) AS FLOAT) - :amt) < 1
                OR ABS(CAST(ISNULL(InvoiceValue, 0) AS FLOAT) - :amt) < 1
              )
            ORDER BY InvoiceID DESC
            """
        ),
        {"cid": int(customer_id), "dt": invoice_date, "amt": amt},
    ).mappings().first()
    return dict(row) if row else None


def _find_existing_invoice(bill_no: str, customer_id, inv_date: date | None) -> dict | None:
    from app.services.gst_invoice_service import GstInvoiceService

    return GstInvoiceService().find_invoice_for_tally_bill(bill_no)


def _build_payload(excel_row: dict, followup: dict, item: dict | None, bank: dict, kind: str) -> dict:
    # DayBook Amount is GST-inclusive invoice value.
    gross = excel_row.get("excel_amount")
    if gross is None:
        gross = followup.get("bill_amount") or 0
    gst_rate = (item or {}).get("gst_rate_percent")
    if gst_rate in (None, ""):
        gst_rate = 18
    intra = _is_intra(followup.get("customer_id"))
    taxable, round_off = _split_inclusive(gross, gst_rate, intra=intra)
    inv_date = excel_row.get("excel_date")
    if inv_date is None and followup.get("invoice_date"):
        inv_date = date.fromisoformat(str(followup["invoice_date"])[:10])
    if inv_date is None:
        inv_date = date.today()
    return {
        "invoice_date": inv_date.isoformat(),
        "invoice_kind": kind,
        "voucher_type": "SALE",
        "customer_id": followup.get("customer_id"),
        "customer_name": followup.get("customer_name") or excel_row.get("excel_name") or "",
        "tally_bill_no": excel_row["bill_no"],
        "payment_bank_account_id": bank.get("account_id"),
        "round_off_amount": str(abs(round_off)),
        "round_off_sign": _round_sign(round_off),
        "lines": [
            {
                "item_id": item.get("item_id") if item else "",
                "tax_period": followup.get("tax_period") or "",
                "quarter": _map_quarter(followup.get("quarter") or ""),
                "month": "",
                "particulars": followup.get("particulars") or f"{followup.get('module_title') or 'Followup'} Followup",
                "hsn_sac": (item or {}).get("hsn_sac") or "",
                "unit": (item or {}).get("unit") or "NOS",
                "qty": 1,
                "rate": float(taxable),
                "discount_amount": 0,
                "gst_rate_percent": gst_rate,
            }
        ],
    }


def _payload_from_existing(record: dict, excel_row: dict) -> dict:
    lines_in = record.get("lines") or []
    line = lines_in[0] if lines_in else {}
    gst_rate = line.get("gst_rate_percent")
    if gst_rate in (None, ""):
        gst_rate = 18
    intra = (record.get("tax_type") or "") == "CGST_SGST"
    gross = excel_row.get("excel_amount") or 0
    taxable, round_off = _split_inclusive(gross, gst_rate, intra=intra)
    new_lines = []
    if lines_in:
        for idx, ln in enumerate(lines_in):
            new_lines.append(
                {
                    "item_id": ln.get("item_id") or "",
                    "tax_period": ln.get("tax_period") or "",
                    "quarter": ln.get("quarter") or "",
                    "month": ln.get("month") or "",
                    "particulars": ln.get("particulars") or "",
                    "hsn_sac": ln.get("hsn_sac") or "",
                    "unit": ln.get("unit") or "NOS",
                    "qty": ln.get("qty") or 1,
                    "rate": float(taxable) if idx == 0 else ln.get("rate") or 0,
                    "discount_amount": 0 if idx == 0 else ln.get("discount_amount") or 0,
                    "gst_rate_percent": gst_rate if idx == 0 else ln.get("gst_rate_percent") or 18,
                }
            )
    else:
        new_lines.append(
            {
                "item_id": "",
                "particulars": "Followup",
                "unit": "NOS",
                "qty": 1,
                "rate": float(taxable),
                "discount_amount": 0,
                "gst_rate_percent": gst_rate,
            }
        )
    return {
        "invoice_date": str(record.get("invoice_date") or "")[:10],
        "invoice_kind": record.get("invoice_kind") or "NON_GST",
        "voucher_type": record.get("voucher_type") or "SALE",
        "customer_id": record.get("customer_id"),
        "customer_name": record.get("customer_name") or "",
        "contact_person": record.get("contact_person") or "",
        "customer_gstin": record.get("customer_gstin") or "",
        "billing_address": record.get("billing_address") or "",
        "contact_mobile": record.get("contact_mobile") or "",
        "contact_email": record.get("contact_email") or "",
        "place_of_supply": record.get("place_of_supply") or "",
        "place_of_supply_code": record.get("place_of_supply_code") or "",
        "reverse_charge": "1" if record.get("reverse_charge") else "0",
        "tally_bill_no": excel_row["bill_no"],
        "payment_bank_account_id": record.get("payment_bank_account_id"),
        "round_off_amount": str(abs(round_off)),
        "round_off_sign": _round_sign(round_off),
        "lines": new_lines,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import DayBook Tally bills as invoices.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to DayBook.xlsx")
    parser.add_argument("--apply", action="store_true", help="Create/update invoices (default is dry-run).")
    parser.add_argument("--fix", action="store_true", help="Fix existing invoices so Invoice Value matches Excel (GST inclusive).")
    parser.add_argument("--limit", type=int, default=0, help="Process only first N unique bill numbers.")
    parser.add_argument("--bank-id", type=int, default=0, help="Payment bank account id (UPI).")
    parser.add_argument("--kind", default="NON_GST", help="GST or NON_GST invoice series.")
    parser.add_argument("--log", default=str(DEFAULT_LOG), help="CSV result path.")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Excel nahi mili: {xlsx_path}")
        return 2

    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
    from app import create_app
    from app.services.followup_service import lookup_tally_bill
    from app.services.gst_invoice_service import GstInvoiceService
    from app.services.item_master_service import ItemMasterService

    rows = _read_sales_rows(xlsx_path)
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    app = create_app()
    results: list[dict] = []
    created = skipped = missing = errors = 0

    with app.app_context():
        inv_svc = GstInvoiceService()
        items = ItemMasterService().list_active_for_dropdown()
        try:
            bank = _pick_bank(args.bank_id or None)
        except ValueError as exc:
            print(str(exc))
            return 3
        print(f"Excel: {xlsx_path}")
        print(f"Unique Tally bills: {len(rows)}")
        print(f"Payment Bank: {bank.get('label')} (id={bank.get('account_id')})")
        if args.fix:
            print(f"Mode: {'FIX-APPLY' if args.apply else 'FIX-DRY-RUN'} (Excel amount = Invoice Value incl. GST)")
            print("-" * 72)
            for row in rows:
                bill_no = row["bill_no"]
                rec = {
                    "bill_no": bill_no,
                    "excel_name": row["excel_name"],
                    "excel_date": row["excel_date"].isoformat() if row["excel_date"] else "",
                    "excel_amount": row["excel_amount"],
                    "status": "",
                    "detail": "",
                }
                followup = lookup_tally_bill(bill_no)
                cid = followup.get("customer_id") if followup else None
                existing = _find_existing_invoice(bill_no, cid, row.get("excel_date"))
                if not existing:
                    missing += 1
                    rec["status"] = "NO_INVOICE"
                    rec["detail"] = "Invoice nahi mili is Tally Bill Number ki"
                    print(f"MISS  {bill_no:>6}  {row['excel_name']}")
                    results.append(rec)
                    continue
                try:
                    record = inv_svc.get_record(int(existing["invoice_id"]))
                    payload = _payload_from_existing(record, row)
                    old_val = float(record.get("invoice_value") or 0)
                    target = float(row.get("excel_amount") or 0)
                    ro_amt = payload.get("round_off_amount") or "0"
                    ro_sign = payload.get("round_off_sign") or ""
                    detail = (
                        f"{record.get('invoice_no')}  {old_val:.2f} -> {target:.2f}  "
                        f"taxable {float(payload['lines'][0]['rate']):.2f}  "
                        f"round {ro_sign} {ro_amt}"
                    )
                    if args.apply:
                        saved = inv_svc.update_record(int(existing["invoice_id"]), payload)
                        created += 1
                        rec["status"] = "FIXED"
                        rec["detail"] = (
                            f"{saved.get('invoice_no')}  Invoice Rs.{float(saved.get('invoice_value') or 0):.2f}  "
                            f"round {ro_sign} {ro_amt}"
                        )
                        print(f"OK    {bill_no:>6}  {rec['detail']}")
                    else:
                        created += 1
                        rec["status"] = "WOULD_FIX"
                        rec["detail"] = detail
                        print(f"READY {bill_no:>6}  {detail}")
                except Exception as exc:
                    errors += 1
                    rec["status"] = "ERROR"
                    rec["detail"] = str(exc)
                    print(f"ERR   {bill_no:>6}  {exc}")
                results.append(rec)
        else:
            print(f"Mode: {'CREATE' if args.apply else 'DRY-RUN'} (Excel amount = Invoice Value incl. GST)")
            print("-" * 72)
            for row in rows:
                bill_no = row["bill_no"]
                rec = {
                    "bill_no": bill_no,
                    "excel_name": row["excel_name"],
                    "excel_date": row["excel_date"].isoformat() if row["excel_date"] else "",
                    "excel_amount": row["excel_amount"],
                    "status": "",
                    "detail": "",
                }
                existing = inv_svc.find_invoice_for_tally_bill(bill_no)
                if existing:
                    skipped += 1
                    rec["status"] = "SKIP_DUPLICATE"
                    rec["detail"] = (
                        f"Invoice already exists: {existing.get('invoice_no')} "
                        f"({existing.get('customer_name')})"
                    )
                    print(f"SKIP  {bill_no:>6}  {rec['detail']}")
                    results.append(rec)
                    continue
                followup = lookup_tally_bill(bill_no)
                if not followup:
                    missing += 1
                    rec["status"] = "NOT_IN_FOLLOWUP"
                    rec["detail"] = "GST/TDS/DSC/ITR Followup mein Tally Bill Number nahi mila"
                    print(f"MISS  {bill_no:>6}  {row['excel_name']}")
                    results.append(rec)
                    continue
                item = _find_item(items, followup.get("module_code") or "")
                payload = _build_payload(row, followup, item, bank, args.kind)
                if not payload.get("customer_id"):
                    errors += 1
                    rec["status"] = "ERROR"
                    rec["detail"] = "Followup customer_id missing"
                    print(f"ERR   {bill_no:>6}  {rec['detail']}")
                    results.append(rec)
                    continue
                similar = _existing_similar_invoice(
                    payload.get("customer_id"),
                    date.fromisoformat(payload["invoice_date"]),
                    row.get("excel_amount"),
                )
                if similar:
                    skipped += 1
                    rec["status"] = "SKIP_DUPLICATE"
                    rec["detail"] = (
                        f"Same customer/date invoice already exists: {similar.get('InvoiceNo')}"
                    )
                    print(f"SKIP  {bill_no:>6}  {rec['detail']}")
                    results.append(rec)
                    continue
                if args.apply:
                    try:
                        saved = inv_svc.create_record(payload, created_by="DayBook Import")
                        created += 1
                        rec["status"] = "CREATED"
                        rec["detail"] = (
                            f"{saved.get('invoice_no')}  {saved.get('customer_name')}  "
                            f"Rs.{float(saved.get('invoice_value') or 0):.2f}  "
                            f"{followup.get('module_code')}"
                        )
                        print(f"OK    {bill_no:>6}  {rec['detail']}")
                    except Exception as exc:
                        errors += 1
                        rec["status"] = "ERROR"
                        rec["detail"] = str(exc)
                        print(f"ERR   {bill_no:>6}  {exc}")
                else:
                    created += 1
                    rec["status"] = "WOULD_CREATE"
                    rec["detail"] = (
                        f"{followup.get('customer_name')}  "
                        f"Invoice Rs.{float(row.get('excel_amount') or 0):.2f}  "
                        f"taxable {float(payload['lines'][0]['rate']):.2f}  "
                        f"round {payload.get('round_off_sign') or '-'} {payload.get('round_off_amount')}  "
                        f"{followup.get('module_code')}"
                    )
                    print(f"READY {bill_no:>6}  {rec['detail']}")
                results.append(rec)

    log_path = Path(args.log)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["bill_no", "excel_name", "excel_date", "excel_amount", "status", "detail"],
        )
        writer.writeheader()
        writer.writerows(results)

    print("-" * 72)
    print(
        f"{'Fixed' if args.fix else 'Ready/Created'}={created}  "
        f"Duplicate={skipped}  Missing={missing}  Error={errors}"
    )
    print(f"Log: {log_path}")
    if not args.apply:
        print("Ye DRY-RUN tha. Apply karne ke liye --apply use karein.")
    return 0 if errors == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
